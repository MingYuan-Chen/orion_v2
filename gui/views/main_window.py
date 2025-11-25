from PySide6.QtWidgets import QMainWindow, QHeaderView, QTableWidgetItem, QInputDialog, QLineEdit
from PySide6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFileDialog, QComboBox, QTabWidget
from PySide6.QtCore import Qt, QTimer, Signal, Slot, QObject, QEvent
from PySide6.QtUiTools import QUiLoader
from typing import Dict, Optional, List
import datetime
from util.logger import logger
from PySide6.QtGui import QColor, QIcon, QFont
import os
import sys
import csv
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from PySide6.QtCore import QFile
from core.services.hardware_test_manager import HardwareTestManagerService
from PySide6.QtWidgets import QApplication
from PySide6.QtWidgets import QScrollArea, QWidget, QVBoxLayout, QSizePolicy, QMessageBox, QFrame, QSpacerItem, QGroupBox

from gui.views.test_manager import TestManagerView
from gui.widgets.test_container import TestContainer
from gui.views.system_info_manager import SystemInfoManagerView
from gui.views.log_manager import LogManagerView
from gui.views.auto_diagnostic_view import AutoDiagnosticView
from gui.views.hw_sw_config_manager import HWSWConfigManager
from gui.views.firmware_os_manager import FirmwareOSManager
from core.services.connection_pre_check import ConnectionPreCheckService
from core.services.cpu_stress_service import CpuStressService
from gui.widgets.simple_cpu_chart_widget import SimpleCpuChartWidget


class DarkEditDialog(QDialog):
    """Custom dark theme edit dialog"""
    
    def __init__(self, parent=None, title="Edit", label_text="Enter value:", initial_text=""):
        super().__init__(parent)
        
        # Set window properties
        self.setWindowTitle(title)
        self.resize(350, 120)
        
        # Set dark theme style
        self.setStyleSheet("""
            QDialog {
                background-color: #2E2E2E;
                color: white;
            }
            QLabel {
                color: white;
                font-weight: bold;
            }
            QLineEdit {
                background-color: #3E3E3E;
                color: #4FC3F7;
                border: 1px solid #555555;
                border-radius: 3px;
                padding: 5px;
                selection-background-color: #0078D7;
            }
            QPushButton {
                background-color: #0078D7;
                color: white;
                border: none;
                padding: 6px 15px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #1C97EA;
            }
            QPushButton:pressed {
                background-color: #00559F;
            }
        """)
        
        # Create layout
        main_layout = QVBoxLayout(self)
        
        # Add label with bold font
        self.label = QLabel(label_text)
        font = self.label.font()
        font.setBold(True)
        self.label.setFont(font)
        main_layout.addWidget(self.label)
        
        # Add line edit
        self.line_edit = QLineEdit(initial_text)
        self.line_edit.selectAll()  # Select all text for easy editing
        main_layout.addWidget(self.line_edit)
        
        # Add buttons
        button_layout = QHBoxLayout()
        
        # Add space to push buttons to the right
        button_layout.addStretch()
        
        # Cancel button
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_button)
        
        # OK button
        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        self.ok_button.setDefault(True)  # Make it the default button (Enter key)
        button_layout.addWidget(self.ok_button)
        
        main_layout.addLayout(button_layout)
        
        # Set focus to the line edit
        self.line_edit.setFocus()
    
    def get_text(self):
        """Get the entered text"""
        return self.line_edit.text()

class MainWindowController(QObject):
    """
    Controller class for managing device monitoring main window
    Each device will create a separate instance
    """
    # Add window closed signal
    window_closed = Signal(str)  # Send device ID
    platform_mapping = {
        "argo": "Argo",
        "hydra_fhd": "Hydra FHD",
        "hydra": "Hydra",
        "gemini_fhd": "Gemini FHD",
        "gemini": "Gemini",
        "athena": "Athena",
        "odin": "Odin"
    }
    
    def __init__(self, device_id, view_model, platform_name=None):
        """
        Initialize main window controller
        
        Args:
            device_id: device ID
            view_model: DeviceManagerViewModel instance
            platform_name: platform name for display in window title
        """
        # Call QObject initialization
        super().__init__()
        
        # Save device ID, view model and platform name
        self.device_id = device_id
        self.view_model = view_model
        self.platform_name = self.platform_mapping.get(platform_name, platform_name.title() if platform_name else "Unknown")
        
        # Add update status flag
        self.is_updating = False
        
        # Add test running status flag
        self.is_test_running = False
        
        # Add USB deployment status flag
        self.usb_deployment_in_progress = False
        
        # add current tab index
        self.current_tab_index = -1
        
        # Add logged commands set to avoid duplicate logging
        self.logged_commands = set()
        
        # Add a dictionary to store the latest system information for export
        self.system_info_data = {}
        
        # Load UI
        try:
            # Get UI file path - support PyInstaller
            if hasattr(sys, '_MEIPASS'):
                # PyInstaller creates a temp folder and stores path in _MEIPASS
                base_path = sys._MEIPASS
                ui_file_path = os.path.join(base_path, 'gui', 'ui', 'main_window.ui')
                icon_path = os.path.join(base_path, 'resources', 'icons', 'header.ico')
            else:
                # Normal development environment
                current_dir = os.path.dirname(os.path.abspath(__file__))
                gui_dir = os.path.dirname(current_dir)
                ui_file_path = os.path.join(gui_dir, "ui", "main_window.ui")
                
                # Get icon path - go up two levels from gui/views
                base_dir = os.path.dirname(gui_dir)
                icon_path = os.path.join(base_dir, "resources", "icons", "header.ico")
            
            logger.debug(f"Loading main window UI from: {ui_file_path}")
            
            # Load UI file
            ui_file = QFile(ui_file_path)
            if not ui_file.open(QFile.ReadOnly):
                error_msg = f"Cannot open {ui_file_path}: {ui_file.errorString()}"
                logger.error(error_msg)
                raise RuntimeError(error_msg)
            
            # Load UI using QUiLoader
            loader = QUiLoader()
            self.window = loader.load(ui_file)
            ui_file.close()
            
            if not self.window:
                error_msg = f"Failed to load UI file: {loader.errorString()}"
                logger.error(error_msg)
                raise RuntimeError(error_msg)
            
            # Ensure the window can be resized
            self.window.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            self.window.setMinimumSize(600, 500)  # set a reasonable minimum size
            self.window.setMaximumSize(16777215, 16777215)  # set the maximum size to a large value
            
            logger.debug(f"Main window UI loaded successfully for device {device_id}")
        except Exception as e:
            logger.error(f"Failed to load main window UI: {str(e)}")
            raise
        
        # create the waiting spinner widget
        from gui.widgets.waiting_spinner import WaitingSpinner
        self.waiting_spinner = WaitingSpinner(
            parent=self.window,
            disable_parent_when_spinning=False,
            radius=10,
            line_length=5,
            line_width=2,
            speed=1.5,
            color=QColor(0, 120, 215)  # use blue, match the system style
        )
        
        # Set window properties
        self._set_window_properties()
        
        # Update dashboard title based on platform name
        self._update_dashboard_title()
        
        # create the system info manager
        self.system_info_manager = SystemInfoManagerView(self.device_id, self.view_model.system_info_service)
        
        # create the HW/SW configuration manager
        self.hw_sw_config_manager = HWSWConfigManager()
        
        # create the firmware & OS manager
        self.firmware_os_manager = FirmwareOSManager()
        
        # create the log manager
        self.log_manager = LogManagerView(self.device_id)
        
        # create the connection pre-check service
        self.connection_pre_check = ConnectionPreCheckService(self.view_model)
        
        # create the CPU stress test service
        self.cpu_stress_service = CpuStressService(self.view_model._serial_worker)
        
        # Stress test state
        self.cpu_stress_duration = 0.0  # Duration in hours
        self.cpu_stress_target_load = 100  # Target CPU load percentage
        self.ram_stress_target_percent = 100  # Target RAM stress percentage
        self.cpu_stress_enabled = True  # CPU stress test enabled
        self.ram_stress_enabled = False  # RAM stress test enabled
        self.cpu_stress_chart_widget = None
        self.device_total_memory_mb = 0  # Device total memory in MB (updated from system info)
        
        # Initialize system info view
        self._init_system_info_view()
        
        # Initialize HW/SW config view
        self._init_hw_sw_config_view()
        
        # Initialize firmware & OS view
        self._init_firmware_os_view()
        
        # Initialize dashboard scrolling
        self._init_dashboard_scrolling()
        
        # Initialize logs view
        self._init_logs_view()
        
        # Create the test manager view and auto diagnostic view
        self.test_manager = TestManagerView(self.device_id, self.view_model.hardware_test_manager, self.platform_name)
        self.auto_diagnostic_view = AutoDiagnosticView(self.device_id, self.view_model.hardware_test_manager)
        
        # set the reference to the test manager
        self.test_manager.main_window_controller = self
        self.auto_diagnostic_view.main_window_controller = self
        
        # Connect signals and slots
        self._connect_signals()
        
        # Initialize auto diagnostic view
        self._init_functionality_test_ui()
        
        # Initialize stress test UI
        self._init_stress_test_ui()
        
        # Connect refresh button
        self.window.pushButton_refresh.clicked.connect(self._on_refresh_system_info)
        
        # Connect USB package deployment signals
        self._connect_usb_deployment_signals()
        
        # Initialize display status as "Initializing"
        self._set_initializing_state()
        
        # Install event filter to capture window close event
        self.window.installEventFilter(self)
        
        # Connect edit button signals
        self.window.button_edit_model_name.clicked.connect(self._on_edit_model_name)
        self.window.button_edit_serial_number.clicked.connect(self._on_edit_serial_number)
        self.window.button_edit_battery_model.clicked.connect(self._on_edit_battery_model)
        self.window.button_edit_battery_serial.clicked.connect(self._on_edit_battery_serial)
        
        # Connect stability test buttons
        self.window.pushButton_battery_monitor.clicked.connect(self._on_battery_monitor_clicked)
        
        # Initialize embedded battery monitor early to ensure refresh button works
        try:
            self._init_embedded_battery_monitor()
            logger.info("Embedded battery monitor initialized during startup")
        except Exception as e:
            logger.warning(f"Failed to initialize embedded battery monitor during startup: {str(e)}")
            # Don't raise the exception here, let it be handled when user clicks the button
        
        # connect the tab changed signal
        if hasattr(self.window, 'tabWidget'):
            self.window.tabWidget.currentChanged.connect(self._on_tab_changed)
        
        # unified test result storage
        self.unified_test_results = {
            "functionality": {},  # functionality test results
            "diagnostic": {}      # diagnostic test results
        }
        self.unified_test_progress = {
            "functionality": {},  # functionality test progress
            "diagnostic": {}      # diagnostic test progress
        }
        
        # Add step template storage to save the step definition for each test
        self.test_step_templates = {
            "functionality": {},  # functionality test step templates
            "diagnostic": {}      # diagnostic test step templates
        }
        
        # set the result recorders after TestManagerView and AutoDiagnosticView are created
        self.test_manager.set_result_recorders(
            lambda test_type, test_id, data: self.record_test_result(test_type, test_id, data),
            lambda test_type, test_id, data: self.record_test_progress(test_type, test_id, data)
        )

        self.auto_diagnostic_view.set_result_recorders(
            lambda test_type, test_id, data: self.record_test_result(test_type, test_id, data),
            lambda test_type, test_id, data: self.record_test_progress(test_type, test_id, data)
        )
        
        # Connect test started signal to save step templates
        self.view_model.hardware_test_manager.test_started.connect(self._on_test_started)

        # System info update will be triggered after USB deployment is completed
        # No need for initial auto-refresh to avoid conflicts with USB deployment commands
    
    @Slot(dict)
    def _on_system_info_updated_main(self, system_info: dict):
        """
        Slot to receive system info updates from SystemInfoManagerView.
        
        Args:
            system_info: A dictionary containing the latest system information.
        """
        self.system_info_data = system_info
        logger.debug(f"Main window received and stored system info update: {list(system_info.keys())}")

    def _on_initial_system_info_refresh(self):
        """initial system info refresh (after the main window is loaded)"""
        # check if the update is already in progress, avoid duplicate execution
        if hasattr(self, 'is_updating') and self.is_updating:
            logger.debug("System info update already in progress, skipping initial refresh")
            return
            
        # Check if USB deployment is in progress, avoid command conflicts
        if hasattr(self, 'usb_deployment_in_progress') and self.usb_deployment_in_progress:
            logger.debug("USB deployment in progress, skipping initial system info refresh")
            return
            
        # use the connection pre-check to ensure the device connection is normal
        self._on_refresh_system_info()
    
    def eventFilter(self, obj, event):
        """Filter window events to capture close event"""
        if obj is self.window and event.type() == QEvent.Close:
            logger.info(f"Main window for device {self.device_id} is closing")
            
            # Stop any ongoing operations
            self.is_updating = False
            self.usb_deployment_in_progress = False
            
            # Stop waiting spinner
            if hasattr(self, 'waiting_spinner'):
                self.waiting_spinner.stop()
                
            # Stop any ongoing system info update
            if hasattr(self.view_model, 'system_info_service') and self.view_model.system_info_service:
                try:
                    self.view_model.system_info_service.stop_update(self.device_id)
                except Exception as e:
                    logger.debug(f"Error stopping system info service: {e}")
            
            # Disconnect system info manager signals
            if hasattr(self, 'system_info_manager') and self.system_info_manager:
                try:
                    self.system_info_manager._disconnect_signals()
                except Exception as e:
                    logger.debug(f"Error disconnecting system info manager signals: {e}")
            
            # Disconnect USB deployment signals
            if hasattr(self, 'view_model') and self.view_model:
                try:
                    self.view_model.usb_deployment_started.disconnect(self._on_usb_deployment_started)
                    self.view_model.usb_deployment_progress.disconnect(self._on_usb_deployment_progress)
                    self.view_model.usb_deployment_completed.disconnect(self._on_usb_deployment_completed)
                    self.view_model.usb_deployment_ready_for_system_info.disconnect(self._on_usb_deployment_ready_for_system_info)
                    logger.debug("Disconnected USB deployment signals")
                except Exception as e:
                    logger.debug(f"Error disconnecting USB deployment signals: {e}")
            
            # Emit window close signal
            self.window_closed.emit(self.device_id)
        return super().eventFilter(obj, event)
    
    def _init_system_info_view(self):
        """Initialize system info view settings"""
        # set the ui components of the system info manager
        system_ui_components = {
            "refresh_button": self.window.pushButton_refresh,
            "last_updated_label": self.window.label_last_updated,
            
            # System basic info
            "model_name": self.window.value_model_name,
            "serial_number": self.window.value_serial_number,
            "cpu": self.window.value_cpu,
            "memory": self.window.value_memory,
            "storage": self.window.value_storage,
            
            # Battery info
            "battery_model": self.window.value_battery_model,
            "battery_serial": self.window.value_battery_serial,
            "voltage": self.window.value_voltage,
            "current": self.window.value_current,
            "design_voltage": self.window.value_design_voltage,
            "design_capacity": self.window.value_design_capacity,
        }
        
        # set the ui components
        self.system_info_manager.set_ui_components(system_ui_components)
        
        # Provide a reference to the controller for command logging
        self.system_info_manager.main_controller = self
        
        # set a function to add system log, directly write to the log manager
        # use anonymous function to wrap, ensure the log will not cause the switch to the system log tab
        self.system_info_manager.add_system_log = lambda level, message: self._add_system_log_without_tab_switch(level, message)
        
        # connect the signals of the system info manager
        # when updating, directly call the method to handle instead of using the signal, so that the tab switch can be controlled
        # when the update is completed, re-enable the UI controls
        self.system_info_manager.info_update_completed.connect(self._on_system_info_update_completed)
        
        # when the update is error, add the error log, and close the waiting icon
        self.system_info_manager.info_update_error.connect(self._on_system_info_update_error)
        
        # Connect the new signal to store system info data for export
        self.system_info_manager.system_info_updated.connect(self._on_system_info_updated_main)
    
    def _init_hw_sw_config_view(self):
        """Initialize HW/SW configuration view"""
        # Get the table widget from the UI
        hw_sw_table = self.window.tableWidget_hw_sw_config
        
        # Set platform name to the HW/SW config manager
        platform_name = getattr(self.view_model, 'platform_name', 'argo')
        self.hw_sw_config_manager.set_platform_name(platform_name)
        
        # Set up the HW/SW config manager with the table widget and edit dialog
        self.hw_sw_config_manager.set_ui_components(hw_sw_table, DarkEditDialog)
        
        # Connect signals if needed
        self.hw_sw_config_manager.config_updated.connect(self._on_hw_sw_config_updated)
        
        logger.info(f"HW/SW configuration view initialized for platform: {platform_name}")
    
    def _init_firmware_os_view(self):
        """Initialize firmware & OS view"""
        # Set up the firmware & OS manager with the window and edit dialog
        self.firmware_os_manager.set_ui_components(self.window, DarkEditDialog)
        
        # Connect signals
        self.firmware_os_manager.info_updated.connect(self._on_firmware_os_updated)
        
        logger.info("Firmware & OS view initialized")
    
    def _init_dashboard_scrolling(self):
        """Initialize dashboard scrolling functionality"""
        try:
            # Check if scroll area already exists in the UI file
            if hasattr(self.window, 'scrollArea_dashboard'):
                logger.info("Scroll area already exists in UI file, configuring existing components")
                
                # Just configure the fixed heights for existing components
                system_overview = self.window.groupBox_system_overview
                if system_overview:
                    system_overview.setMinimumHeight(380)
                    system_overview.setMaximumHeight(380)
                    logger.debug("Set System Overview fixed height: 380px")
                
                # Find and configure HW Components
                hw_components = self.window.tableWidget_hw_sw_config
                if hw_components:
                    # Find the parent groupbox
                    hw_groupbox = hw_components.parent()
                    while hw_groupbox and not isinstance(hw_groupbox, QGroupBox):
                        hw_groupbox = hw_groupbox.parent()
                    
                    if hw_groupbox:
                        hw_groupbox.setMinimumHeight(400)
                        hw_groupbox.setMaximumHeight(400)
                        logger.debug("Set HW Components group fixed height: 400px")
                    
                    # Set the table to show all rows without internal scrolling
                    hw_components.setMinimumHeight(350)
                    hw_components.setMaximumHeight(350)
                    hw_components.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                    logger.debug("Set HW Components table fixed height: 350px, no internal scrolling")
                
                # Find and configure Firmware & OS group
                firmware_os_groupbox = self.window.groupBox_firmware_os
                if firmware_os_groupbox:
                    firmware_os_groupbox.setMinimumHeight(200)
                    firmware_os_groupbox.setMaximumHeight(200)
                    logger.debug("Set Firmware & OS group fixed height: 200px")
                
                logger.info("Dashboard scrolling functionality configured successfully")
                return
            
            # If no scroll area exists, create one dynamically (fallback)
            logger.info("No scroll area found in UI file, creating dynamically")
            
            # Get the dashboard tab
            dashboard_tab = self.window.tab_dashboard
            
            # Get the existing layout
            existing_layout = dashboard_tab.layout()
            if not existing_layout:
                logger.warning("No existing layout found in dashboard tab")
                return
            
            # Create a scroll area
            scroll_area = QScrollArea(dashboard_tab)
            scroll_area.setFrameShape(QScrollArea.NoFrame)
            scroll_area.setWidgetResizable(True)
            scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            
            # Create content widget for scroll area
            content_widget = QWidget()
            content_layout = QVBoxLayout(content_widget)
            content_layout.setContentsMargins(10, 10, 10, 10)
            content_layout.setSpacing(10)
            
            # Move all existing widgets to the content layout
            while existing_layout.count():
                item = existing_layout.takeAt(0)
                if item.widget():
                    content_layout.addWidget(item.widget())
                elif item.layout():
                    content_layout.addLayout(item.layout())
                elif item.spacerItem():
                    content_layout.addItem(item.spacerItem())
            
            # Set fixed heights for groups
            system_overview = self.window.groupBox_system_overview
            if system_overview:
                system_overview.setMinimumHeight(580)
                system_overview.setMaximumHeight(580)
                logger.debug("Set System Overview fixed height: 580px")
            
            # Add a vertical spacer at the bottom to handle extra space
            spacer = QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding)
            content_layout.addItem(spacer)
            
            # Set the content widget to the scroll area
            scroll_area.setWidget(content_widget)
            
            # Remove the old layout and add the scroll area
            dashboard_tab.setLayout(None)
            new_layout = QVBoxLayout(dashboard_tab)
            new_layout.setContentsMargins(0, 0, 0, 0)
            new_layout.addWidget(scroll_area)
            
            logger.info("Dashboard scrolling functionality created dynamically")
            
        except Exception as e:
            logger.error(f"Error initializing dashboard scrolling: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
    
    def _on_hw_sw_config_updated(self, component_id: str, field_type: str, new_value: str):
        """Handle HW/SW configuration update"""
        # Log the configuration change
        self.log_manager.add_log_entry(
            "INFO", 
            f"HW/SW Config updated - {component_id} {field_type}: {new_value}"
        )
        logger.info(f"HW/SW Config updated: {component_id} {field_type} = {new_value}")
    
    def _on_firmware_os_updated(self, field_name: str, new_value: str):
        """Handle firmware & OS information update"""
        # Log the firmware & OS change
        self.log_manager.add_log_entry(
            "INFO", 
            f"Firmware & OS updated - {field_name}: {new_value}"
        )
        logger.info(f"Firmware & OS updated: {field_name} = {new_value}")

    def _add_system_log_without_tab_switch(self, level, message):
        """Add system log without switching tabs"""
        # record the current tab
        current_tab = -1
        if hasattr(self.window, 'tabWidget'):
            current_tab = self.window.tabWidget.currentIndex()
            
        # add the log, but do not scroll to bottom to avoid activating the log tab
        self.log_manager.add_log_entry(level, message, scroll_to_bottom=False)
        
        # if updating and current tab is valid, restore to the previous tab
        if self.is_updating and current_tab >= 0 and hasattr(self.window, 'tabWidget'):
            self.window.tabWidget.setCurrentIndex(current_tab)
    
    def _on_system_info_update_completed(self):
        """Handle the event of system info update completed"""
        # restore the updating status flag
        self.is_updating = False
        
        # stop the waiting icon
        if hasattr(self, 'waiting_spinner'):
            self.waiting_spinner.stop()
        
        # Update device memory information for RAM stress test
        self._update_device_memory_info()
        
        # Use the unified test execution completion mechanism
        self._on_test_execution_completed("System Info Refresh")
        
        # add the completed log
        self.log_manager.add_log_entry("INFO", "System info update completed")
    
    def _on_system_info_update_error(self, error_message):
        """Handle the event of system info update error"""
        # restore the updating status flag
        self.is_updating = False
        
        # stop the waiting icon
        if hasattr(self, 'waiting_spinner'):
            self.waiting_spinner.stop()
        
        # Use the unified test execution abort mechanism
        self._on_test_execution_aborted("System Info Refresh", f"Error: {error_message}")
        
        # add the error log
        self.log_manager.add_log_entry("ERROR", f"System info update error: {error_message}")
    
    def _init_logs_view(self):
        """Initialize logs view settings"""
        # Configure logs table
        logs_table = self.window.tableWidget_logs
        logs_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        logs_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)  # Message column stretch
        
        # Set column width
        logs_table.setColumnWidth(0, 180)  # Timestamp column
        logs_table.setColumnWidth(1, 80)   # Level column
        
        # Enable automatic row height adjustment to display full content
        logs_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        
        # Set table to automatically wrap text
        logs_table.setWordWrap(True)
        
        # set the ui components to the log manager
        self.log_manager.set_ui_components(
            self.window.tableWidget_logs,
            self.window.comboBox_log_level,
            self.window.comboBox_time_range,
            self.window.pushButton_clear_logs
        )
        
        # Connect command send button
        self.window.pushButton_send_command.clicked.connect(self._on_send_command)
        self.window.lineEdit_command.returnPressed.connect(self._on_send_command)
        
        # Connect logs refresh button
        self.window.pushButton_refresh_logs.clicked.connect(self._refresh_logs)
        
        # Connect CTRL+C button
        self.window.pushButton_send_ctrl_c.clicked.connect(self._on_send_ctrl_c)
    
    def _filter_logs(self):
        """Filter logs based on selected level and time range"""
        # this method is already handled internally by LogManagerView, keep it for backward compatibility
        pass
    
    def _refresh_logs(self):
        """Refresh log data"""
        # In actual application, get latest logs from device
        logger.info(f"Refreshing logs for device: {self.device_id}")
        
        # For example, can send command to get logs
        self.view_model.send_command(self.device_id, "get_logs")
        
        # Can update logs in command_completed signal processing
    
    def _clear_logs(self):
        """Clear log table"""
        # delegate to the log manager to clear logs
        self.log_manager.clear_logs()
    
    def _on_send_command(self):
        """Process command sending"""
        command = self.window.lineEdit_command.text().strip()
        if not command:
            return
            
        logger.info(f"Sending command to device {self.device_id}: {command}")
        
        # Send command to device
        self.view_model.send_command(self.device_id, command)
        
        # Clear command input box
        self.window.lineEdit_command.clear()
        
        # Add command record to logs
        self.log_manager.add_log_entry("INFO", f"[Command] {command}")
        
        # mark the command as logged
        self.mark_command_as_logged(command)
    
    def _on_send_ctrl_c(self):
        """Process CTRL+C sending"""
        logger.info(f"Sending CTRL+C to device {self.device_id}")
        
        # Send CTRL+C to device
        self.view_model.send_ctrl_c(self.device_id)
        
        # Add CTRL+C record to logs
        self.log_manager.add_log_entry("INFO", "[CTRL+C] Interrupt signal sent")
    
    def _connect_signals(self):
        """Connect signals and slots"""
        # Connect command result signal from view model
        self.view_model.command_result.connect(self._on_command_completed)
        
        # connect the all tests completed signal of the test manager
        self.test_manager.all_tests_completed.connect(self._on_all_tests_completed)
        
        # Connect CPU stress test signals
        self.cpu_stress_service.stress_started.connect(self._on_cpu_stress_started)
        self.cpu_stress_service.stress_completed.connect(self._on_cpu_stress_completed)
        self.cpu_stress_service.stress_error.connect(self._on_cpu_stress_error)
        self.cpu_stress_service.stress_progress.connect(self._on_cpu_stress_progress)
        self.cpu_stress_service.temp_warning.connect(self._on_temp_warning)
        
        # CPU stress service manages its own command result connection internally

    def _init_functionality_test_ui(self):
        """Initialize functionality test UI elements"""
        # Configure shared test steps table
        hw_test_table = self.window.tableWidget_hardware_test_steps
        hw_test_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        hw_test_table.horizontalHeader().setStretchLastSection(True)
        hw_test_table.setColumnWidth(0, 200)  # Step column
        hw_test_table.setColumnWidth(1, 80)   # Status column
        
        # Set table minimum height
        hw_test_table.verticalHeader().setDefaultSectionSize(30)  # Set row height to 30 pixels
        hw_test_table.setMinimumHeight(180)   # Set minimum height
        
        # get the functionality test tab
        tab_functionality = self.window.tab_functionality
        
        # set the background color of the tab
        tab_functionality.setStyleSheet("background-color: #1E1E1E;")
        
        # clear all the child widgets in the tab
        for child in tab_functionality.findChildren(QWidget):
            # skip the already processed UI components
            if child.objectName() in ["tableWidget_hardware_test_steps", "progressBar_hardware_test"]:
                continue
            # skip the button widgets
            if isinstance(child, QPushButton) and child in [self.window.button_test_all]:
                continue
            # only process the direct child widgets
            if child.parent() == tab_functionality:
                child.setVisible(False)
                child.deleteLater()
        
        # save the original layout reference
        original_layout = tab_functionality.layout()
        if original_layout:
            # remove all the elements in the existing layout (avoid duplicate addition)
            while original_layout.count():
                item = original_layout.takeAt(0)
                if item.widget():
                    item.widget().setParent(None)
        else:
            # if there is no layout, create a new one
            original_layout = QVBoxLayout(tab_functionality)
            original_layout.setContentsMargins(5, 5, 5, 5)
        
        # create the scroll area
        scroll_area = QScrollArea(tab_functionality)
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QScrollArea.NoFrame)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        
        # set the scroll bar style
        scroll_area.setStyleSheet("""
            QScrollArea {
                background-color: #1E1E1E;
                border: none;
            }
            QScrollBar:vertical {
                background: #333333;
                width: 10px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #555555;
                border-radius: 4px;
                min-height: 20px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
        """)
        
        # create the content container
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(5, 5, 5, 5)
        content_layout.setSpacing(10)
        
        # set the background color of the content container
        content_widget.setStyleSheet("background-color: #1E1E1E;")
        
        # create the auto diagnostic element
        self.auto_diagnostic_widget = self.auto_diagnostic_view.create_widget()
        content_layout.addWidget(self.auto_diagnostic_widget)
        
        # set the auto diagnostic test items
        diagnostic_tests = {}
        diagnostic_tests["diagnostic_cpu_name"] = "Check CPU Name"
        diagnostic_tests["diagnostic_cpu_processor"] = "Check CPU Processor"
        diagnostic_tests["diagnostic_design_capacity"] = "Check Design Capacity"
        diagnostic_tests["diagnostic_design_voltage"] = "Check Design Voltage"
        diagnostic_tests["diagnostic_emmc_size"] = "Check eMMC Size"
        diagnostic_tests["diagnostic_kernal_name"] = "Check Kernal Name"
        diagnostic_tests["diagnostic_mac_address"] = "Check MAC Address"
        diagnostic_tests["diagnostic_memory_size"] = "Check Memory Size"
        diagnostic_tests["diagnostic_nor_flash_size"] = "Check NOR Flash Size"
        diagnostic_tests["diagnostic_panel_id"] = "Check Panel ID"
        diagnostic_tests["diagnostic_panel_resolution"] = "Check Panel Resolution"
        diagnostic_tests["diagnostic_pic_version"] = "Check PIC Version"
        diagnostic_tests["diagnostic_sync_time"] = "Check Sync Time"
        diagnostic_tests["diagnostic_uboot_version"] = "Check U-Boot Version"
        diagnostic_tests["diagnostic_wifi_bt"] = "Check Wifi and Bluetooth"
        # diagnostic_tests["diagnostic_set_get_rtc_time"] = "Check Set and Get RTC Time"
        
        if self.platform_name == "Athena":
            diagnostic_tests["diagnostic_ethernet"] = "Check Ethernet Connection"
            diagnostic_tests["diagnostic_wifi_connection"] = "Check Wifi Connection"
            diagnostic_tests.pop("diagnostic_nor_flash_size")
            diagnostic_tests.pop("diagnostic_panel_id")
        if self.platform_name == "Odin":
            diagnostic_tests.pop("diagnostic_nor_flash_size")
            diagnostic_tests.pop("diagnostic_panel_id")

        self.auto_diagnostic_view.setup_diagnostic_items(diagnostic_tests)
        
        # set the auto diagnostic log function
        self.auto_diagnostic_view.add_system_log = lambda level, message: self.log_manager.add_log_entry(level, message)
        
        # connect the auto diagnostic signal
        self.auto_diagnostic_view.all_diagnostics_completed.connect(self._on_all_diagnostics_completed)
        self.auto_diagnostic_view.export_report_requested.connect(self._export_results)
        
        # add the spacing
        spacer = QSpacerItem(20, 30, QSizePolicy.Minimum, QSizePolicy.Fixed)
        content_layout.addItem(spacer)
        
        # create the functionality test title element
        hw_title_widget = QWidget()
        hw_title_layout = QHBoxLayout(hw_title_widget)
        hw_title_layout.setContentsMargins(5, 5, 5, 5)
        
        # set the background color of the title area
        hw_title_widget.setStyleSheet("background-color: #1E1E1E;")
        
        hw_title = QLabel("Functionality Test")
        hw_title.setStyleSheet("font-weight: bold; color: #4FC3F7; font-size: 14px;")
        hw_title_layout.addWidget(hw_title)
        
        hw_title_layout.addStretch()
        
        # add the test all button
        if hasattr(self.window, 'button_test_all'):
            # if the button_test_all does not exist, create a new one
            if not self.window.button_test_all:
                self.window.button_test_all = QPushButton("Test All")
            
            # apply the style
            self.window.button_test_all.setStyleSheet("""
                QPushButton {
                    background-color: #0078D7;
                    color: white;
                    border: none;
                    padding: 4px 12px;
                    border-radius: 3px;
                }
                QPushButton:hover {
                    background-color: #1C97EA;
                }
            """)
            hw_title_layout.addWidget(self.window.button_test_all)
        else:
            # if there is no this attribute, create a new button
            self.window.button_test_all = QPushButton("Test All")
            self.window.button_test_all.setStyleSheet("""
                QPushButton {
                    background-color: #0078D7;
                    color: white;
                    border: none;
                    padding: 4px 12px;
                    border-radius: 3px;
                }
                QPushButton:hover {
                    background-color: #1C97EA;
                }
            """)
            hw_title_layout.addWidget(self.window.button_test_all)
        
        # create the abort button
        self.window.button_abort_test = QPushButton("Abort Test")
        self.window.button_abort_test.setStyleSheet("""
            QPushButton {
                background-color: #D32F2F;
                color: white;
                border: none;
                padding: 4px 12px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #F44336;
            }
        """)
        hw_title_layout.addWidget(self.window.button_abort_test)
        
        # add the title to the layout
        content_layout.addWidget(hw_title_widget)
        
        # add the separator
        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setFrameShadow(QFrame.Sunken)
        separator.setStyleSheet("background-color: #333333;")
        separator.setMaximumHeight(1)
        content_layout.addWidget(separator)
        
        # create the test container
        test_container = TestContainer()
        
        # set the fixed height of the test container, to display more items
        item_height = 32  # the height of each test item
        visible_items = 5  # the number of items to display, same as Auto Diagnostic
        scroll_height = item_height * visible_items + 20  # add some extra space
        test_container.set_fixed_height(scroll_height)
        
        content_layout.addWidget(test_container)
        
        # add the spacing
        progress_spacer = QSpacerItem(20, 30, QSizePolicy.Minimum, QSizePolicy.Fixed)
        content_layout.addItem(progress_spacer)
        
        # add the test progress title
        progress_title = QLabel("Test Progress")
        progress_title.setStyleSheet("font-weight: bold; color: white; font-size: 13px; background-color: #1E1E1E;")
        content_layout.addWidget(progress_title)
        
        # add the test steps table
        content_layout.addWidget(hw_test_table)
        
        # set the background color of the table
        hw_test_table.setStyleSheet("""
            QTableWidget {
                background-color: #1E1E1E;
                color: white;
                gridline-color: #333333;
                border: none;
            }
            QTableWidget::item {
                background-color: #252526;
            }
            QHeaderView::section {
                background-color: #252526;
                color: white;
                border: 1px solid #333333;
                padding: 4px;
            }
        """)
        
        # add the progress bar
        if hasattr(self.window, 'progressBar_hardware_test'):
            content_layout.addWidget(self.window.progressBar_hardware_test)
        
        # add the bottom spacing
        bottom_spacer = QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Fixed)
        content_layout.addItem(bottom_spacer)
        
        # set the content of the scroll area
        scroll_area.setWidget(content_widget)
        
        # add the scroll area to the original layout
        original_layout.addWidget(scroll_area)
        
        # ensure the entire tab page has the correct background color
        tab_functionality.update()
        
        # find and clear any unknown text labels at the bottom of the tab page (excluding known UI components)
        for label in tab_functionality.findChildren(QLabel):
            if label.parent() == tab_functionality:  # only process direct child labels
                # skip title and button UI components
                if label.text() in ["Functionality Test", "Test Progress", "Auto Diagnostic"]:
                    continue
                # only process labels without objectName or obviously redundant labels
                if not label.objectName() or label.objectName() == "":
                    label.setVisible(False)
                    label.deleteLater()
        
        # set the UI components first, so we can register tests
        self.test_manager.set_ui_components(
            test_container,
            self.window.button_test_all,
            self.window.tableWidget_hardware_test_steps,
            self.window.progressBar_hardware_test,
            self.window,
            self.window.button_abort_test
        )

        # Register the test items through the test manager
        self.test_manager.register_test("functionality_audio", "Audio Test")
        self.test_manager.register_test("functionality_backlight", "Backlight Test")
        self.test_manager.register_test("functionality_battery", "Battery Test")
        if self.platform_name != "Odin":
            self.test_manager.register_test("functionality_camera", "Camera Test")
        charge_test_name = "Charge Test" if self.platform_name != "Athena" else "Charge Setting Test"
        logger.debug(f"Charge test name set as {charge_test_name}")
        self.test_manager.register_test("functionality_charge", charge_test_name)
        self.test_manager.register_test("functionality_eeprom", "EEPROM Test")
        self.test_manager.register_test("functionality_emmc", "eMMC Test")
        if self.platform_name != "Odin":
            self.test_manager.register_test("functionality_hdmi", "HDMI Test")
        self.test_manager.register_test("functionality_lcd", "LCD Test")
        self.test_manager.register_test("functionality_led", "LED Test")
        self.test_manager.register_test("functionality_power_button", "Power Button Test")
        self.test_manager.register_test("functionality_touch", "Touch Test")
        self.test_manager.register_test("functionality_usb", "USB Test")
        if self.platform_name not in ["Athena", "Argo", "Hydra", "Gemini"]:
            self.test_manager.register_test("functionality_probe", "Probe Test")
        
        # set the log recorder
        self.test_manager.add_system_log = lambda level, message: self.log_manager.add_log_entry(level, message)
        
        # set the default hidden progress bar
        self.window.progressBar_hardware_test.setVisible(False)
        
        # set the result recorder
        self.test_manager.set_result_recorders(
            lambda test_type, test_id, data: self.record_test_result(test_type, test_id, data),
            lambda test_type, test_id, data: self.record_test_progress(test_type, test_id, data)
        )

        self.auto_diagnostic_view.set_result_recorders(
            lambda test_type, test_id, data: self.record_test_result(test_type, test_id, data),
            lambda test_type, test_id, data: self.record_test_progress(test_type, test_id, data)
        )
    
    def _init_stress_test_ui(self):
        """Initialize stress test UI elements"""
        try:
            # Set default CPU loading to 100%
            cpu_loading_combo = self.window.comboBox_cpu_loading
            cpu_loading_combo.setCurrentIndex(9)  # 100% is at index 9
            
            # Set default RAM loading to 100%
            ram_loading_spinbox = self.window.spinBox_ram_loading
            ram_loading_spinbox.setValue(100)
            
            # Set default duration to 0.0
            duration_spinbox = self.window.doubleSpinBox_duration
            duration_spinbox.setValue(0.0)
            
            # Set default checkbox states
            self.window.checkBox_cpu_stress.setChecked(True)   # CPU enabled by default
            self.window.checkBox_ram_stress.setChecked(False)  # RAM disabled by default
            
            # Connect UI signals
            self.window.doubleSpinBox_duration.valueChanged.connect(self._on_duration_changed)
            self.window.pushButton_cpu_stress_start.clicked.connect(self._on_cpu_stress_start_clicked)
            self.window.comboBox_cpu_loading.currentTextChanged.connect(self._on_cpu_loading_changed)
            self.window.spinBox_ram_loading.valueChanged.connect(self._on_ram_loading_changed)
            self.window.checkBox_cpu_stress.toggled.connect(self._on_cpu_stress_toggled)
            self.window.checkBox_ram_stress.toggled.connect(self._on_ram_stress_toggled)
            
            # Set scroll area style for stress test tab
            self._setup_stress_test_scroll_area()
            
            # Initialize CPU stress chart widget
            self._init_cpu_stress_chart()
            
            # Update UI state based on checkbox selection
            self._update_stress_test_ui_state()
            
            logger.debug("Stress test UI initialized successfully")
            
        except Exception as e:
            logger.error(f"Error initializing stress test UI: {e}")
    
    def _setup_stress_test_scroll_area(self):
        """Setup scroll area style for stress test tab"""
        try:
            scroll_area = self.window.scrollArea_stress_test
            scroll_area.setStyleSheet("""
                QScrollArea {
                    background-color: #1E1E1E;
                    border: none;
                }
                QScrollBar:vertical {
                    background: #333333;
                    width: 12px;
                    margin: 0px;
                    border-radius: 6px;
                }
                QScrollBar::handle:vertical {
                    background: #555555;
                    border-radius: 6px;
                    min-height: 20px;
                    margin: 2px;
                }
                QScrollBar::handle:vertical:hover {
                    background: #666666;
                }
                QScrollBar::handle:vertical:pressed {
                    background: #777777;
                }
                QScrollBar::add-line:vertical, 
                QScrollBar::sub-line:vertical {
                    height: 0px;
                    background: none;
                }
            """)
            logger.debug("Stress test scroll area style applied")
            
        except Exception as e:
            logger.error(f"Error setting up stress test scroll area: {e}")
    
    def _init_cpu_stress_chart(self):
        """Initialize CPU stress chart widget"""
        try:
            # Create CPU stress chart widget
            self.cpu_stress_chart_widget = SimpleCpuChartWidget()
            
            # Get the placeholder widget and replace it with the chart
            placeholder = self.window.widget_cpu_chart_placeholder
            parent_layout = placeholder.parent().layout()
            
            # Find the placeholder in the layout and replace it
            for i in range(parent_layout.count()):
                item = parent_layout.itemAt(i)
                if item.widget() == placeholder:
                    # Remove the placeholder
                    parent_layout.removeItem(item)
                    placeholder.deleteLater()
                    
                    # Add the chart widget
                    parent_layout.insertWidget(i, self.cpu_stress_chart_widget)
                    break
            
            # 連接溫度閾值變化信號
            self.cpu_stress_chart_widget.warning_temp_spinbox.valueChanged.connect(
                self._on_warning_temp_threshold_changed
            )
            
            # 設置設備 ID 用於 CSV 檔案命名
            if hasattr(self, 'device_id') and self.device_id:
                self.cpu_stress_chart_widget.set_device_id(self.device_id)
            else:
                self.cpu_stress_chart_widget.set_device_id("unknown")
            
            logger.debug("CPU stress chart initialized successfully")
            
        except Exception as e:
            logger.error(f"Error initializing CPU stress chart: {e}")
    
    def _on_duration_changed(self, value: float):
        """Handle duration spinbox value change"""
        self.cpu_stress_duration = value
        logger.debug(f"CPU stress duration changed to: {value} hours")
    
    def _on_cpu_loading_changed(self, text: str):
        """Handle CPU loading selection change"""
        try:
            self.cpu_stress_target_load = int(text.replace('%', ''))
            logger.debug(f"CPU target load changed to: {self.cpu_stress_target_load}%")
            
        except ValueError:
            self.cpu_stress_target_load = 100
    
    def _on_ram_loading_changed(self, value: int):
        """Handle RAM loading spinbox value change"""
        self.ram_stress_target_percent = value
        
        # Update tooltip with actual MB value and safety warnings
        if self.device_total_memory_mb > 0:
            target_mb = int((value / 100.0) * self.device_total_memory_mb)
            
            # Generate tooltip with safety information
            tooltip_text = f"{value}% = {target_mb} MB (Total memory: {self.device_total_memory_mb} MB)"
            
            # Add safety warnings for high percentages
            if value >= 90:
                tooltip_text += "\n⚠️ Warning: High RAM usage may affect system stability"
            elif value >= 80:
                tooltip_text += "\n⚠️ Warning: High RAM usage may affect system response"
            elif value >= 50:
                tooltip_text += "\n💡 Tip: Medium RAM stress test"
            else:
                tooltip_text += "\n✅ Safe: Low RAM stress test"
                
            self.window.spinBox_ram_loading.setToolTip(tooltip_text)
            logger.debug(f"RAM target load changed to: {value}% ({target_mb} MB of {self.device_total_memory_mb} MB total)")
        else:
            self.window.spinBox_ram_loading.setToolTip("Device memory size unknown")
            logger.debug(f"RAM target load changed to: {value}% (total memory unknown)")
    
    def _on_cpu_stress_toggled(self, checked: bool):
        """Handle CPU stress checkbox toggle"""
        self.cpu_stress_enabled = checked
        self._update_stress_test_ui_state()
        logger.debug(f"CPU stress test {'enabled' if checked else 'disabled'}")
    
    def _on_ram_stress_toggled(self, checked: bool):
        """Handle RAM stress checkbox toggle"""
        self.ram_stress_enabled = checked
        self._update_stress_test_ui_state()
        logger.debug(f"RAM stress test {'enabled' if checked else 'disabled'}")
    
    def _update_stress_test_ui_state(self):
        """Update stress test UI controls based on checkbox states"""
        try:
            # Enable/disable CPU controls based on CPU checkbox
            cpu_enabled = self.window.checkBox_cpu_stress.isChecked()
            self.window.comboBox_cpu_loading.setEnabled(cpu_enabled)
            
            # Enable/disable RAM controls based on RAM checkbox  
            ram_enabled = self.window.checkBox_ram_stress.isChecked()
            self.window.spinBox_ram_loading.setEnabled(ram_enabled)
            
            # Ensure at least one stress type is selected
            if not cpu_enabled and not ram_enabled:
                self.window.pushButton_cpu_stress_start.setEnabled(False)
                self.window.pushButton_cpu_stress_start.setToolTip("Select at least one stress test type")
            else:
                self.window.pushButton_cpu_stress_start.setEnabled(True)
                self.window.pushButton_cpu_stress_start.setToolTip("")
            
        except Exception as e:
            logger.error(f"Error updating stress test UI state: {e}")
    
    def _get_device_total_memory_mb(self) -> int:
        """Get device total memory in MB from system info UI display"""
        try:
            # Try to get memory info from the UI label that shows memory information
            if hasattr(self.window, 'value_memory'):
                memory_text = self.window.value_memory.text()  # e.g., "3.8Gi (180Mi Used)"
                
                if memory_text and memory_text != "":
                    # Extract total memory from the display text
                    # Format: "3.8Gi (180Mi Used)" or just "3.8Gi"
                    memory_str = memory_text.split('(')[0].strip()  # Get "3.8Gi" part
                    
                    # Parse memory string to MB
                    memory_str = memory_str.strip().upper()
                    if memory_str.endswith('GI') or memory_str.endswith('G'):
                        # Remove suffix and convert GB to MB
                        value = float(memory_str.replace('GI', '').replace('G', ''))
                        return int(value * 1024)  # 1 GB = 1024 MB
                    elif memory_str.endswith('MI') or memory_str.endswith('M'):
                        # Already in MB
                        value = float(memory_str.replace('MI', '').replace('M', ''))
                        return int(value)
                    elif memory_str.endswith('KI') or memory_str.endswith('K'):
                        # Convert KB to MB
                        value = float(memory_str.replace('KI', '').replace('K', ''))
                        return int(value / 1024)  # 1024 KB = 1 MB
                        
        except Exception as e:
            logger.warning(f"Error parsing device memory info from UI: {e}")
        
        return 0
    
    def _update_device_memory_info(self):
        """Update device memory information and adjust RAM stress UI accordingly"""
        try:
            self.device_total_memory_mb = self._get_device_total_memory_mb()
            if self.device_total_memory_mb > 0:
                logger.info(f"Device total memory: {self.device_total_memory_mb} MB")
                
                # Update the tooltip to show actual MB value with safety information
                if hasattr(self.window, 'spinBox_ram_loading'):
                    current_percent = self.window.spinBox_ram_loading.value()
                    actual_mb = int((current_percent / 100.0) * self.device_total_memory_mb)
                    
                    # Generate tooltip with safety information
                    tooltip_text = f"{current_percent}% = {actual_mb} MB (Total memory: {self.device_total_memory_mb} MB)"
                    
                    # Add safety warnings for high percentages
                    if current_percent >= 90:
                        tooltip_text += "\n⚠️ Warning: High RAM usage may affect system stability"
                    elif current_percent >= 80:
                        tooltip_text += "\n⚠️ Warning: High RAM usage may affect system response"
                    elif current_percent >= 50:
                        tooltip_text += "\n💡 Tip: Medium RAM stress test"
                    else:
                        tooltip_text += "\n✅ Safe: Low RAM stress test"
                        
                    self.window.spinBox_ram_loading.setToolTip(tooltip_text)
                    
                # 更新圖表組件的總記憶體信息
                if self.cpu_stress_chart_widget:
                    self.cpu_stress_chart_widget.set_total_ram_mb(self.device_total_memory_mb)
            else:
                logger.warning("Could not determine device total memory")
                if hasattr(self.window, 'spinBox_ram_loading'):
                    self.window.spinBox_ram_loading.setToolTip("Device memory size unknown")
                    
        except Exception as e:
            logger.error(f"Error updating device memory info: {e}")
    
    def _on_cpu_stress_start_clicked(self):
        """Handle stress test start/stop button click"""
        try:
            button = self.window.pushButton_cpu_stress_start
            
            if button.text() == "Start":
                # Get current settings
                self.cpu_stress_duration = self.window.doubleSpinBox_duration.value()
                cpu_enabled = self.window.checkBox_cpu_stress.isChecked()
                ram_enabled = self.window.checkBox_ram_stress.isChecked()
                
                # Validate inputs
                if self.cpu_stress_duration < 0:
                    self.add_system_log("ERROR", "Duration cannot be negative")
                    return
                
                if not cpu_enabled and not ram_enabled:
                    self.add_system_log("ERROR", "Select at least one stress test type")
                    return
                
                # Safety check for high RAM usage
                if ram_enabled and self.ram_stress_target_percent >= 90:
                    from PySide6.QtWidgets import QMessageBox
                    
                    # Create custom message box with dark styling
                    msg_box = QMessageBox(self.window)
                    msg_box.setWindowTitle("High RAM usage warning")
                    msg_box.setText(f"You are about to use {self.ram_stress_target_percent}% of system memory for stress testing.\n\n"
                                  f"This may:\n"
                                  f"• Affect system stability\n"
                                  f"• Cause system response slowdown\n"
                                  f"• In extreme cases, may cause system reboot\n\n"
                                  f"Are you sure you want to continue?")
                    msg_box.setIcon(QMessageBox.Warning)
                    msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                    msg_box.setDefaultButton(QMessageBox.No)
                    
                    # Apply dark styling
                    msg_box.setStyleSheet(self._get_dark_message_box_style())
                    
                    reply = msg_box.exec()
                    if reply == QMessageBox.No:
                        self.add_system_log("INFO", f"User cancelled {self.ram_stress_target_percent}% RAM stress test")
                        return
                
                # Convert hours to seconds (0 means unlimited duration)
                if self.cpu_stress_duration == 0:
                    duration_seconds = 0  # 0 means unlimited
                    duration_text = "unlimited time"
                else:
                    duration_seconds = int(self.cpu_stress_duration * 3600)
                    duration_text = f"{self.cpu_stress_duration} hours"
                
                # Determine stress test parameters
                cpu_loading = self.cpu_stress_target_load if cpu_enabled else 0
                
                # Calculate RAM MB from percentage
                ram_mb = 0
                if ram_enabled:
                    if self.device_total_memory_mb > 0:
                        ram_mb = int((self.ram_stress_target_percent / 100.0) * self.device_total_memory_mb)
                        # Ensure minimum of 64MB
                        ram_mb = max(64, ram_mb)
                    else:
                        # Fallback: if device memory unknown, use 256MB for 25%
                        ram_mb = int(self.ram_stress_target_percent * 10.24)  # Approximate for ~4GB system
                        ram_mb = max(64, ram_mb)
                        logger.warning(f"Device memory unknown, using estimated {ram_mb}MB for {self.ram_stress_target_percent}%")
                
                # Create description of stress test
                stress_types = []
                if cpu_enabled:
                    stress_types.append(f"CPU {cpu_loading}%")
                if ram_enabled:
                    stress_types.append(f"RAM {self.ram_stress_target_percent}% ({ram_mb}MB)")
                stress_description = " + ".join(stress_types)
                
                # Start the stress test
                success = self.cpu_stress_service.start_stress_test(
                    self.device_id, 
                    cpu_loading, 
                    duration_seconds,
                    ram_mb
                )
                
                if success:
                    button.setText("Stop")
                    button.setStyleSheet("background-color: #D7342A;")  # Red for stop
                    
                    # Disable controls during test
                    self.window.checkBox_cpu_stress.setEnabled(False)
                    self.window.checkBox_ram_stress.setEnabled(False)
                    self.window.comboBox_cpu_loading.setEnabled(False)
                    self.window.spinBox_ram_loading.setEnabled(False)
                    self.window.doubleSpinBox_duration.setEnabled(False)
                    
                    # 自動開始 CSV 記錄
                    if hasattr(self, 'cpu_stress_chart_widget') and self.cpu_stress_chart_widget:
                        self.cpu_stress_chart_widget.start_csv_logging()
                    
                    self.add_system_log("INFO", f"Stress test started: {stress_description} for {duration_text}")
                
            else:
                # Stop stress test
                success = self.cpu_stress_service.stop_stress_test(self.device_id)
                if success:
                    # 自動停止 CSV 記錄
                    if hasattr(self, 'cpu_stress_chart_widget') and self.cpu_stress_chart_widget:
                        self.cpu_stress_chart_widget.stop_csv_logging()
                    
                    self._reset_cpu_stress_ui()
                    self.add_system_log("INFO", "Stress test stopped manually")
                    
        except Exception as e:
            logger.error(f"Error handling stress test start/stop: {e}")
            self.add_system_log("ERROR", f"Stress test error: {e}")
    
    def _reset_cpu_stress_ui(self):
        """Reset stress test UI to initial state"""
        try:
            button = self.window.pushButton_cpu_stress_start
            button.setText("Start")
            button.setStyleSheet("")  # Reset to default style
            
            # Re-enable controls
            self.window.checkBox_cpu_stress.setEnabled(True)
            self.window.checkBox_ram_stress.setEnabled(True)
            self.window.doubleSpinBox_duration.setEnabled(True)
            
            # Update UI state based on current checkbox selections
            self._update_stress_test_ui_state()
            
        except Exception as e:
            logger.error(f"Error resetting stress test UI: {e}")
    
    @Slot(str, int, int)
    def _on_cpu_stress_started(self, device_id: str, loading_percent: int, duration_seconds: int):
        """Handle CPU stress test started signal"""
        if device_id == self.device_id:
            duration_hours = duration_seconds / 3600.0
            self.add_system_log("INFO", f"CPU stress test started: {loading_percent}% load for {duration_hours:.1f} hours")
            
            # Clear chart data for new test
            if self.cpu_stress_chart_widget:
                self.cpu_stress_chart_widget.clear_data()
                self.cpu_stress_chart_widget.set_chart_title(f"CPU Temperature Monitor - Stress: {loading_percent}%")
    
    @Slot(str, str)
    def _on_cpu_stress_completed(self, device_id: str, message: str):
        """Handle CPU stress test completed signal"""
        if device_id == self.device_id:
            # 自動停止 CSV 記錄
            if hasattr(self, 'cpu_stress_chart_widget') and self.cpu_stress_chart_widget:
                self.cpu_stress_chart_widget.stop_csv_logging()
            
            self._reset_cpu_stress_ui()
            self.add_system_log("INFO", f"CPU stress test completed: {message}")
    
    @Slot(str, str)
    def _on_cpu_stress_error(self, device_id: str, error_message: str):
        """Handle CPU stress test error signal"""
        if device_id == self.device_id:
            # 自動停止 CSV 記錄
            if hasattr(self, 'cpu_stress_chart_widget') and self.cpu_stress_chart_widget:
                self.cpu_stress_chart_widget.stop_csv_logging()
            
            self._reset_cpu_stress_ui()
            self.add_system_log("ERROR", f"CPU stress test error: {error_message}")
    
    @Slot(str, int, int, int, int, float)
    def _on_cpu_stress_progress(self, device_id: str, elapsed_seconds: int, total_seconds: int, ram_stress_enabled: int, ram_stress_mb: int, cpu_temp: float):
        """Handle CPU stress test progress signal"""
        if device_id == self.device_id:
            # Use target CPU load as display value (consistent with stress test setting)
            # Display the configured stress level instead of simulated random variation
            simulated_load = self.cpu_stress_target_load
            
            # Add data point to chart with temperature as main data
            if self.cpu_stress_chart_widget:
                self.cpu_stress_chart_widget.add_data_point(
                    cpu_temp,  # Mainly display temperature
                    simulated_load,  # CPU load for logging
                    bool(ram_stress_enabled), 
                    ram_stress_mb
                )
            
            # Log progress periodically (every 30 seconds)
            if elapsed_seconds % 30 == 0:
                if total_seconds == 0:  # Unlimited duration
                    self.add_system_log("INFO", f"CPU stress test progress: {elapsed_seconds}s (unlimited duration)")
                else:
                    progress_percent = (elapsed_seconds / total_seconds) * 100
                    self.add_system_log("INFO", f"CPU stress test progress: {elapsed_seconds}/{total_seconds}s ({progress_percent:.1f}%)")
    
    @Slot(int)
    def _on_warning_temp_threshold_changed(self, threshold: int):
        """Handle warning temperature threshold change"""
        # 更新 CPU stress service 的閾值
        self.cpu_stress_service.set_temperature_warning_threshold(float(threshold))
        logger.debug(f"Updated CPU stress service warning threshold to: {threshold}°C")
    
    @Slot(str, float)
    def _on_temp_warning(self, device_id: str, temperature: float):
        """Handle CPU temperature warning signal"""
        if device_id == self.device_id:
            # 獲取當前閾值用於動態顯示
            current_threshold = self.cpu_stress_service.temp_warning_threshold
            self.add_system_log("WARNING", f"CPU temperature warning: {temperature:.1f}°C - Above {current_threshold:.0f}°C threshold, hardware thermal protection will activate if needed")
    
    def _set_initializing_state(self):
        """Set all system info display to initializing state"""
        # delegate to the system info manager to set the initializing state
        self.system_info_manager.set_initializing_state()
    
    def _connect_usb_deployment_signals(self):
        """Connect USB package deployment signals"""
        try:
            # Connect USB deployment signals
            self.view_model.usb_deployment_started.connect(self._on_usb_deployment_started)
            self.view_model.usb_deployment_progress.connect(self._on_usb_deployment_progress)
            self.view_model.usb_deployment_completed.connect(self._on_usb_deployment_completed)
            self.view_model.usb_deployment_ready_for_system_info.connect(self._on_usb_deployment_ready_for_system_info)
            
            logger.debug("USB deployment signals connected")
            
        except Exception as e:
            logger.error(f"Error connecting USB deployment signals: {str(e)}")
    
    def _start_usb_package_deployment(self):
        """Start USB package deployment"""
        try:
            logger.info(f"Starting USB package deployment for device {self.device_id}")
            
            # Set USB deployment in progress flag
            self.usb_deployment_in_progress = True
            
            # Update status to show deployment in progress
            self.system_info_manager.set_deployment_status("Deploying USB package...")
            
            # Position and show the waiting icon in the spinner placeholder
            if hasattr(self, 'waiting_spinner'):
                self._position_spinner_in_placeholder()
                self.waiting_spinner.start()
            
            # Start deployment
            self.view_model.start_usb_package_deployment(
                self.device_id,
                on_success=self._on_usb_deployment_success,
                on_failure=self._on_usb_deployment_failure
            )
            
        except Exception as e:
            logger.error(f"Error starting USB package deployment: {str(e)}")
            # Clear deployment flag and proceed with system info update
            self.usb_deployment_in_progress = False
            self._on_refresh_system_info()
    
    def _on_usb_deployment_started(self, device_id: str):
        """Handle USB deployment started signal"""
        if device_id == self.device_id:
            logger.info(f"USB deployment started for device {device_id}")
            self.system_info_manager.set_deployment_status("Scanning USB devices...")
    
    def _on_usb_deployment_progress(self, device_id: str, progress_message: str):
        """Handle USB deployment progress signal"""
        if device_id == self.device_id:
            logger.info(f"USB deployment progress for device {device_id}: {progress_message}")
            self.system_info_manager.set_deployment_status(progress_message)
    
    def _on_usb_deployment_completed(self, device_id: str, success: bool, message: str):
        """Handle USB deployment completed signal"""
        if device_id == self.device_id:
            logger.info(f"USB deployment completed for device {device_id}: success={success}, message={message}")
            
            # Clear USB deployment in progress flag
            self.usb_deployment_in_progress = False
            
            # Stop the waiting spinner
            if hasattr(self, 'waiting_spinner'):
                self.waiting_spinner.stop()
                
            if success:
                self.system_info_manager.set_deployment_status(f"USB package deployment successful: {message}")
            else:
                self.system_info_manager.set_deployment_status(f"USB package deployment failed: {message}")
    
    def _on_usb_deployment_ready_for_system_info(self, device_id: str):
        """Handle USB deployment ready for system info signal"""
        if device_id == self.device_id:
            logger.info(f"USB deployment ready for system info for device {device_id}")
            
            # Check if system info update is already in progress or has been triggered
            if hasattr(self, 'is_updating') and self.is_updating:
                logger.debug("System info update already in progress, ignoring USB deployment ready signal")
                return
                
            # Ensure USB deployment flag is cleared before starting system info
            self.usb_deployment_in_progress = False
            # Now start system info update
            self._on_refresh_system_info()
    
    def _on_usb_deployment_success(self):
        """Handle USB deployment success callback"""
        logger.info(f"USB deployment success callback for device {self.device_id}")
    
    def _on_usb_deployment_failure(self, reason: str):
        """Handle USB deployment failure callback"""
        logger.warning(f"USB deployment failure callback for device {self.device_id}: {reason}")
        
        # Clear USB deployment in progress flag
        self.usb_deployment_in_progress = False
        
        # Stop the waiting spinner
        if hasattr(self, 'waiting_spinner'):
            self.waiting_spinner.stop()
            
        # Even if deployment fails, proceed with system info update
        self._on_refresh_system_info()

    def _position_spinner_in_placeholder(self):
        """Position waiting spinner in the designated placeholder"""
        if hasattr(self.window, 'label_spinner_placeholder'):
            placeholder = self.window.label_spinner_placeholder
            if placeholder and placeholder.isVisible():
                # Get the global position of the placeholder
                global_pos = placeholder.mapToGlobal(placeholder.rect().center())
                # Convert back to parent widget coordinates
                parent_pos = self.waiting_spinner.parent().mapFromGlobal(global_pos)
                # Center the spinner in the placeholder
                spinner_x = parent_pos.x() - self.waiting_spinner.width() // 2
                spinner_y = parent_pos.y() - self.waiting_spinner.height() // 2
                self.waiting_spinner.move(spinner_x, spinner_y)

    def _on_refresh_system_info(self):
        """Handle refresh button click with pre-connection check"""
        # check if the update is already in progress, avoid duplicate execution
        if hasattr(self, 'is_updating') and self.is_updating:
            logger.debug("System info update already in progress, ignoring duplicate request")
            return
            
        # Check if USB deployment is in progress, avoid command conflicts
        if hasattr(self, 'usb_deployment_in_progress') and self.usb_deployment_in_progress:
            logger.debug("USB deployment in progress, postponing system info update")
            return
            
        # Clear USB deployment status and restore normal display
        self.system_info_manager.clear_deployment_status()
            
        # record the current tab, but do not force switch back, allow the user to freely switch
        if hasattr(self.window, 'tabWidget'):
            self.current_tab_index = self.window.tabWidget.currentIndex()
            logger.debug(f"Current tab index: {self.current_tab_index} (Dashboard)")
        
        # add the log, but do not switch to the log tab
        self.log_manager.add_log_entry("INFO", f"Checking connection before refreshing system info for {self.device_id}...")
        
        # position and show the waiting icon in the spinner placeholder
        if hasattr(self, 'waiting_spinner'):
            self._position_spinner_in_placeholder()
            self.waiting_spinner.start()
        
        # use the connection pre-check service to execute the system info refresh
        self.connection_pre_check.execute_with_pre_check(
            device_id=self.device_id,
            operation_name="System Info Refresh",
            operation_callback=self._execute_system_info_refresh,
            on_success=self._on_system_info_pre_check_success,
            on_failure=self._on_system_info_pre_check_failure,
            check_timeout=12000  # 12 seconds timeout
        )
    
    def _execute_system_info_refresh(self):
        """execute the actual system info refresh operation"""
        # Use the unified test execution mechanism
        self._on_test_execution_started("System Info Refresh")
        
        # set the updating status flag
        self.is_updating = True
        
        # execute the system info refresh
        self.system_info_manager.refresh_system_info()
    
    def _on_system_info_pre_check_success(self):
        """system info refresh pre-check success callback"""
        self.log_manager.add_log_entry("INFO", f"Connection verified, starting system info refresh for {self.device_id}")
    
    def _on_system_info_pre_check_failure(self, reason: str):
        """system info refresh pre-check failure callback"""
        self.log_manager.add_log_entry("ERROR", f"Connection check failed for system info refresh: {reason}")
        
        # stop the possible ongoing system info update
        if hasattr(self.view_model, 'system_info_service') and self.view_model.system_info_service:
            self.view_model.system_info_service.stop_update(self.device_id)
        
        # Use the unified test execution abort mechanism
        self._on_test_execution_aborted("System Info Refresh", f"Pre-check failed: {reason}")
        
        # restore the UI state (additional cleanup)
        self.is_updating = False
        
        # stop the waiting animation
        if hasattr(self, 'waiting_spinner'):
            self.waiting_spinner.stop()
        
        # show the error message
        msg_box = QMessageBox(self.window)
        msg_box.setWindowTitle("Connection Check Failed")
        msg_box.setText("Device connection failed, system info refresh is canceled.")
        msg_box.setInformativeText(f"Ensure the device is connected, back up existing test content, then close the main window and return to Device Manager to reconnect the device.")
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setStandardButtons(QMessageBox.Ok)
        
        # apply
        msg_box.setStyleSheet(self._get_dark_message_box_style())
        
        msg_box.exec()
    
    def execute_functionality_test_with_pre_check(self, test_id: str = None):
        """execute the functionality test with pre-check"""
        operation_name = f"Functionality Test ({test_id})" if test_id else "Functionality Test All"
        
        self.log_manager.add_log_entry("INFO", f"Checking connection before starting {operation_name.lower()}...")
        
        # use the connection pre-check service to execute the functionality test
        self.connection_pre_check.execute_with_pre_check(
            device_id=self.device_id,
            operation_name=operation_name,
            operation_callback=lambda: self._execute_functionality_test(test_id),
            on_success=lambda: self._on_functionality_test_pre_check_success(operation_name),
            on_failure=lambda reason: self._on_functionality_test_pre_check_failure(operation_name, reason, test_id),
            check_timeout=12000  # 12 seconds timeout
        )
    
    def _execute_functionality_test(self, test_id: str = None):
        """execute the actual functionality test operation"""
        # Disable UI controls when functionality test starts
        test_name = f"Functionality Test ({test_id})" if test_id else "Functionality Test All"
        self._on_test_execution_started(test_name)
        
        if test_id:
            # execute the single test
            self.test_manager._start_test_directly(test_id)
        else:
            # execute all tests
            self.test_manager._start_test_all_directly()
    
    def _on_functionality_test_pre_check_success(self, operation_name: str):
        """functionality test pre-check success callback"""
        self.log_manager.add_log_entry("INFO", f"Connection verified, starting {operation_name.lower()}")
    
    def _on_functionality_test_pre_check_failure(self, operation_name: str, reason: str, test_id: str = None):
        """functionality test pre-check failure callback"""
        self.log_manager.add_log_entry("ERROR", f"Connection check failed for {operation_name.lower()}: {reason}")
        
        # Restore UI state since pre-check failed
        self._on_test_execution_aborted(operation_name, f"Pre-check failed: {reason}")
        
        # Reset test manager button state since pre-check failed
        if hasattr(self, 'test_manager') and self.test_manager:
            if test_id:
                # Reset individual test button state
                self.test_manager.test_container.set_test_state(test_id, "not_started")
            else:
                # Reset Test All button state
                self.test_manager._reset_test_all_button_state()
        
        # show the error message
        msg_box = QMessageBox(self.window)
        msg_box.setWindowTitle("Connection Check Failed")
        msg_box.setText(f"Device connection failed, {operation_name} is canceled.")
        msg_box.setInformativeText(f"Ensure the device is connected, back up existing test content, then close the main window and return to Device Manager to reconnect the device.")
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setStandardButtons(QMessageBox.Ok)
        
        # apply dark style sheet
        msg_box.setStyleSheet(self._get_dark_message_box_style())
        
        msg_box.exec()
    
    def execute_auto_diagnostic_with_pre_check(self):
        """execute the auto diagnostic with pre-check"""
        self.log_manager.add_log_entry("INFO", "Checking connection before starting auto diagnostic...")
        
        # use the connection pre-check service to execute the auto diagnostic
        self.connection_pre_check.execute_with_pre_check(
            device_id=self.device_id,
            operation_name="Auto Diagnostic",
            operation_callback=self._execute_auto_diagnostic,
            on_success=self._on_auto_diagnostic_pre_check_success,
            on_failure=self._on_auto_diagnostic_pre_check_failure,
            check_timeout=12000  # 12 seconds timeout
        )
    
    def _execute_auto_diagnostic(self):
        """execute the actual auto diagnostic operation"""
        # Disable UI controls when auto diagnostic starts
        self._on_test_execution_started("Auto Diagnostic")
        
        self.auto_diagnostic_view._run_all_tests_directly()
    
    def _on_auto_diagnostic_pre_check_success(self):
        """auto diagnostic pre-check success callback"""
        self.log_manager.add_log_entry("INFO", "Connection verified, starting auto diagnostic")
    
    def _on_auto_diagnostic_pre_check_failure(self, reason: str):
        """auto diagnostic pre-check failure callback"""
        self.log_manager.add_log_entry("ERROR", f"Connection/Login check failed for auto diagnostic: {reason}")
        
        # Restore UI state since pre-check failed
        self._on_test_execution_aborted("Auto Diagnostic", f"Pre-check failed: {reason}")
        
        # Reset auto diagnostic button state since pre-check failed
        if hasattr(self, 'auto_diagnostic_view') and self.auto_diagnostic_view:
            self.auto_diagnostic_view.is_running = False
            if self.auto_diagnostic_view.run_all_button:
                self.auto_diagnostic_view.run_all_button.setText("Run All Tests")
                self.auto_diagnostic_view.run_all_button.setEnabled(True)
        
        # show the error message
        msg_box = QMessageBox(self.window)
        msg_box.setWindowTitle("Connection/Login Check Failed")
        msg_box.setText("Device connection/login failed, auto diagnostic is canceled.")
        msg_box.setInformativeText(f"Ensure the device is connected, back up existing test content, then close the main window and return to Device Manager to reconnect the device.")
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setStandardButtons(QMessageBox.Ok)
        
        # apply dark style sheet
        msg_box.setStyleSheet(self._get_dark_message_box_style())
        
        msg_box.exec()

    def _on_edit_model_name(self):
        """handle the edit model name button click"""
        if self.system_info_manager.edit_model_name():
            self.log_manager.add_log_entry("INFO", "Model name updated")

    def _on_edit_serial_number(self):
        """handle the edit serial number button click"""
        if self.system_info_manager.edit_serial_number():
            self.log_manager.add_log_entry("INFO", "Serial number updated")

    def _on_edit_battery_model(self):
        """handle the edit battery model button click"""
        if self.system_info_manager.edit_battery_model():
            self.log_manager.add_log_entry("INFO", "Battery model updated")

    def _on_edit_battery_serial(self):
        """handle the edit battery serial number button click"""
        if self.system_info_manager.edit_battery_serial():
            self.log_manager.add_log_entry("INFO", "Battery serial number updated")

    def _on_battery_monitor_clicked(self):
        """Handle Battery Monitor button click with pre-connection check"""
        logger.info("Battery Monitor button clicked")
        
        # Add log entry
        self.log_manager.add_log_entry("INFO", "Battery Monitor button clicked")
        
        try:
            # Initialize embedded battery monitor if not already done
            if not hasattr(self, 'battery_monitor_manager') or self.battery_monitor_manager is None:
                logger.info("Battery monitor not initialized, initializing now...")
                self._init_embedded_battery_monitor()
            
            # Check if manager is available and determine current state
            if hasattr(self, 'battery_monitor_manager') and self.battery_monitor_manager:
                is_currently_monitoring = self.battery_monitor_manager.is_monitoring
                logger.info(f"Current monitoring state: {is_currently_monitoring}")
                
                # Check if button is stuck in checking state
                button_text = self.window.pushButton_battery_monitor.text()
                if button_text == "Checking...":
                    logger.warning("Button is stuck in Checking state, forcing reset")
                    self.battery_monitor_manager.force_reset()
                    self._restore_battery_monitor_button_state()
                    self.log_manager.add_log_entry("WARNING", "Reset battery monitor from stuck checking state")
                    return
                
                if is_currently_monitoring:
                    # Immediately show stopping state before calling stop_monitoring
                    logger.info("Setting button to Stopping... state immediately")
                    self.window.pushButton_battery_monitor.setText("Stopping...")
                    self.window.pushButton_battery_monitor.setEnabled(False)
                    self.window.pushButton_battery_monitor.setStyleSheet("""
                        QPushButton {
                            background-color: #FF9800;
                            color: white;
                            border: none;
                            padding: 6px 15px;
                            border-radius: 3px;
                        }
                        QPushButton:disabled {
                            background-color: #FF9800;
                            color: white;
                        }
                    """)
                    
                    # Force immediate UI update
                    self.window.pushButton_battery_monitor.repaint()
                    
                    # Now call stop monitoring
                    self.battery_monitor_manager.stop_monitoring()
                else:
                    # Start monitoring with pre-check
                    self.execute_battery_monitoring_with_pre_check()
            else:
                logger.error("Battery monitor manager not available after initialization")
                
        except Exception as e:
            logger.error(f"Error in battery monitor button click: {str(e)}")
            self.log_manager.add_log_entry("ERROR", f"Battery monitor error: {str(e)}")

    def execute_battery_monitoring_with_pre_check(self):
        """Execute battery monitoring with pre-check"""
        # Check if update is already in progress, avoid duplicate execution
        if hasattr(self, 'is_updating') and self.is_updating:
            logger.debug("System update already in progress, postponing battery monitoring")
            return
            
        # Check if USB deployment is in progress, avoid command conflicts
        if hasattr(self, 'usb_deployment_in_progress') and self.usb_deployment_in_progress:
            logger.debug("USB deployment in progress, postponing battery monitoring")
            return
        
        self.log_manager.add_log_entry("INFO", f"Checking connection before starting battery monitoring for {self.device_id}...")
        
        # Show checking state
        self.window.pushButton_battery_monitor.setText("Checking...")
        self.window.pushButton_battery_monitor.setEnabled(False)
        self.window.pushButton_battery_monitor.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                border: none;
                padding: 6px 15px;
                border-radius: 3px;
            }
            QPushButton:disabled {
                background-color: #FF9800;
                color: white;
            }
        """)
        
        # Use connection pre-check service to execute battery monitoring
        self.connection_pre_check.execute_with_pre_check(
            device_id=self.device_id,
            operation_name="Battery Monitoring",
            operation_callback=self._execute_battery_monitoring,
            on_success=self._on_battery_monitoring_pre_check_success,
            on_failure=self._on_battery_monitoring_pre_check_failure,
            check_timeout=12000  # 12 seconds timeout
        )
    
    def _execute_battery_monitoring(self):
        """Execute the actual battery monitoring operation"""
        try:
            # Start battery monitoring
            if hasattr(self, 'battery_monitor_manager') and self.battery_monitor_manager:
                success = self.battery_monitor_manager.start_monitoring()
                if not success:
                    # If start_monitoring failed, restore button state
                    logger.warning("Battery monitoring failed to start")
                    self._restore_battery_monitor_button_state()
                    self.log_manager.add_log_entry("ERROR", "Failed to start battery monitoring")
                else:
                    logger.info("Battery monitoring started successfully")
            else:
                logger.error("Battery monitor manager not available")
                self._restore_battery_monitor_button_state()
                self.log_manager.add_log_entry("ERROR", "Battery monitor manager not available")
        except Exception as e:
            logger.error(f"Error executing battery monitoring: {str(e)}")
            self._restore_battery_monitor_button_state()
            self.log_manager.add_log_entry("ERROR", f"Battery monitoring error: {str(e)}")
    
    def _restore_battery_monitor_button_state(self):
        """Restore battery monitor button to normal state"""
        self.window.pushButton_battery_monitor.setText("Start Monitoring")
        self.window.pushButton_battery_monitor.setEnabled(True)
        self.window.pushButton_battery_monitor.setStyleSheet("""
            QPushButton {
                background-color: #0078D7;
                color: white;
                border: none;
                padding: 6px 15px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #1C97EA;
            }
        """)
        logger.debug("Battery monitor button state restored")
    
    def _on_battery_monitoring_pre_check_success(self):
        """Battery monitoring pre-check success callback"""
        self.log_manager.add_log_entry("INFO", f"Connection verified, starting battery monitoring for {self.device_id}")
        
        # Force clear any lingering states before starting
        if hasattr(self, 'battery_monitor_manager') and self.battery_monitor_manager:
            if hasattr(self.battery_monitor_manager, 'battery_service') and self.battery_monitor_manager.battery_service:
                self.battery_monitor_manager.battery_service._single_reading_mode = False
                self.battery_monitor_manager.battery_service._is_processing = False
                logger.debug("Pre-start: Cleared battery service state flags")
    
    def _on_battery_monitoring_pre_check_failure(self, reason: str):
        """Battery monitoring pre-check failure callback"""
        self.log_manager.add_log_entry("ERROR", f"Connection check failed for battery monitoring: {reason}")
        
        # Force reset states on failure to ensure clean state
        if hasattr(self, 'battery_monitor_manager') and self.battery_monitor_manager:
            self.battery_monitor_manager.force_reset()
        
        # Restore button to normal state since pre-check failed
        self.window.pushButton_battery_monitor.setText("Start Monitoring")
        self.window.pushButton_battery_monitor.setEnabled(True)
        self.window.pushButton_battery_monitor.setStyleSheet("""
            QPushButton {
                background-color: #0078D7;
                color: white;
                border: none;
                padding: 6px 15px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #1C97EA;
            }
        """)
        
        # Show error message
        msg_box = QMessageBox(self.window)
        msg_box.setWindowTitle("Connection Check Failed")
        msg_box.setText("Device connection failed, battery monitoring is canceled.")
        msg_box.setInformativeText(f"Ensure the device is connected, back up existing test content, then close the main window and return to Device Manager to reconnect the device.")
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setStandardButtons(QMessageBox.Ok)
        
        # Apply dark style sheet
        msg_box.setStyleSheet(self._get_dark_message_box_style())
        
        msg_box.exec()

    def _init_embedded_battery_monitor(self):
        """Initialize embedded battery monitor for main window"""
        
        try:
            # Get serial worker from view model
            serial_worker = None
            if hasattr(self.view_model, '_serial_worker') and self.view_model._serial_worker:
                serial_worker = self.view_model._serial_worker
            
            if serial_worker is None:
                raise Exception("Serial worker not available")
            
            # Get platform name
            platform_name = getattr(self.view_model, 'platform_name', 'argo')
            
            # Import required classes
            from core.services.battery_monitor_service import BatteryMonitorService
            from gui.views.battery_monitor_manager import BatteryMonitorManager
            
            # Create battery monitor service
            self.battery_service = BatteryMonitorService(serial_worker, platform_name)
            
            # Create battery monitor manager
            self.battery_monitor_manager = BatteryMonitorManager(self.device_id, self.battery_service)
            
            # Initialize battery chart widget
            self._init_battery_chart()
            
            # Map UI components from main window (exclude refresh_button to handle separately)
            ui_mapping = {
                "monitor_button": self.window.pushButton_battery_monitor,
                # "refresh_button": self.window.pushButton_battery_refresh,  # Remove this line
                "interval_spinbox": self.window.spinBox_monitor_interval,
                "status_label": self.window.label_battery_status,
                "voltage_label": self.window.label_voltage_value,
                "current_label": self.window.label_current_value,
                "temperature_label": self.window.label_temperature_value,
                "battery_level_label": self.window.label_battery_level_value,
                "progress_bar": self.window.progressBar_battery_level,
                "led_status_label": self.window.label_led_status_value,
                "interrupt_status_label": self.window.label_interrupt_status_value,
                "battery_status_label": self.window.label_dc_status_value,
                "cpu_usage_label": self.window.label_cpu_usage_value,
                "memory_usage_label": self.window.label_memory_usage_value,
                "log_as_file_checkbox": self.window.checkBox_log_as_file
            }
            
            self.battery_monitor_manager.set_ui_components(ui_mapping)
            
            # Set chart widget reference BEFORE setting main controller
            if hasattr(self, 'battery_chart_widget') and self.battery_chart_widget:
                self.battery_monitor_manager.set_chart_widget(self.battery_chart_widget)
                logger.info("Chart widget reference set to battery monitor manager")
            
            # Set main controller reference for battery monitor manager
            self.battery_monitor_manager.main_controller = self
            
            # Connect signals
            self.battery_monitor_manager.monitoring_started.connect(self._on_embedded_monitoring_started)
            self.battery_monitor_manager.monitoring_completed.connect(self._on_embedded_monitoring_completed)
            self.battery_monitor_manager.monitoring_error.connect(self._on_embedded_monitoring_error)
            
            # Connect battery refresh button to our pre-check method
            self.window.pushButton_battery_refresh.clicked.connect(self._on_battery_refresh_clicked)
            
            # Connect battery service signals for proper data flow
            if hasattr(self.battery_service, 'battery_info_received'):
                self.battery_service.battery_info_received.connect(self.battery_monitor_manager._on_battery_info_received)
            if hasattr(self.battery_service, 'battery_info_error'):
                self.battery_service.battery_info_error.connect(self.battery_monitor_manager._on_battery_info_error)
            
            # Connect display control checkboxes to monitor configuration handler
            self._connect_battery_display_controls()
            
            # Don't perform initial refresh to avoid state conflicts
            # User can click "Refresh Once" button if needed
            
            logger.info("Embedded battery monitor initialized")
            
        except Exception as e:
            logger.error(f"Failed to initialize embedded battery monitor: {str(e)}")
            raise
    
    def add_system_log(self, level: str, message: str):
        """
        Add system log entry (for battery monitor manager compatibility)
        
        Args:
            level: Log level (INFO, WARNING, ERROR, etc.)
            message: Log message
        """
        self.log_manager.add_log_entry(level, message)
    
    def _connect_battery_display_controls(self):
        """Connect battery display control checkboxes to monitor configuration handler"""
        try:
            # Map of checkbox names to monitor names
            checkbox_monitor_map = {
                'checkBox_show_battery_level': 'relative_state',
                'checkBox_show_voltage_monitor': 'voltage', 
                'checkBox_show_current_monitor': 'current',
                'checkBox_show_temperature_monitor': 'temperature',
                'checkBox_show_led_status': 'led_status',
                'checkBox_show_cpu_usage': 'cpu_usage',
                'checkBox_show_memory_usage': 'memory_usage',
                'checkBox_show_battery_status': 'battery_status',
                'checkBox_show_interrupt_status': 'interrupt_status'
            }
            
            # Connect each checkbox to the handler
            for checkbox_name, monitor_name in checkbox_monitor_map.items():
                if hasattr(self.window, checkbox_name):
                    checkbox = getattr(self.window, checkbox_name)
                    # Use lambda with default parameter to capture monitor_name correctly
                    checkbox.toggled.connect(
                        lambda checked, name=monitor_name: self._on_battery_monitor_toggle(name, checked)
                    )
                    logger.debug(f"Connected {checkbox_name} to monitor {monitor_name}")
            
            # Set initial monitor configuration based on current checkbox states
            self._update_battery_monitor_configuration()
            
            logger.info("Battery display control checkboxes connected successfully")
            
        except Exception as e:
            logger.error(f"Failed to connect battery display control checkboxes: {str(e)}")
    
    def _on_battery_monitor_toggle(self, monitor_name: str, checked: bool):
        """
        Handle battery monitor checkbox toggle
        
        Args:
            monitor_name: Name of the monitor (e.g., 'voltage', 'current')
            checked: Whether the checkbox is checked
        """
        logger.info(f"Battery monitor toggle: {monitor_name} = {checked}")
        
        # Update the battery service configuration
        if hasattr(self, 'battery_service') and self.battery_service:
            current_config = self.battery_service.get_enabled_monitors()
            current_config[monitor_name] = checked
            self.battery_service.set_enabled_monitors(current_config)
            
            # Add system log
            self.log_manager.add_log_entry("INFO", f"Battery monitor {monitor_name} {'enabled' if checked else 'disabled'}")
        
        # Update UI display visibility based on checkbox state
        self._update_battery_display_visibility(monitor_name, checked)
    
    def _update_battery_monitor_configuration(self):
        """Update battery monitor configuration based on current checkbox states"""
        try:
            if not hasattr(self, 'battery_service') or not self.battery_service:
                return
            
            # Map of checkbox names to monitor names
            checkbox_monitor_map = {
                'checkBox_show_battery_level': 'relative_state',
                'checkBox_show_voltage_monitor': 'voltage',
                'checkBox_show_current_monitor': 'current', 
                'checkBox_show_temperature_monitor': 'temperature',
                'checkBox_show_led_status': 'led_status',
                'checkBox_show_cpu_usage': 'cpu_usage',
                'checkBox_show_memory_usage': 'memory_usage',
                'checkBox_show_battery_status': 'battery_status',
                'checkBox_show_interrupt_status': 'interrupt_status'
            }
            
            # Get current checkbox states
            monitor_config = {}
            for checkbox_name, monitor_name in checkbox_monitor_map.items():
                if hasattr(self.window, checkbox_name):
                    checkbox = getattr(self.window, checkbox_name)
                    monitor_config[monitor_name] = checkbox.isChecked()
            
            # Update battery service configuration
            self.battery_service.set_enabled_monitors(monitor_config)
            
            logger.info(f"Battery monitor configuration updated: {monitor_config}")
            
        except Exception as e:
            logger.error(f"Failed to update battery monitor configuration: {str(e)}")
    
    def _update_battery_display_visibility(self, monitor_name: str, visible: bool):
        """
        Update battery display UI element visibility
        
        Args:
            monitor_name: Name of the monitor
            visible: Whether the display should be visible
        """
        try:
            # Map monitor names to UI elements
            monitor_ui_map = {
                'relative_state': ['label_battery_level_title', 'label_battery_level_value', 'progressBar_battery_level'],
                'voltage': ['label_voltage_title', 'label_voltage_value'],
                'current': ['label_current_title', 'label_current_value'], 
                'temperature': ['label_temperature_title', 'label_temperature_value'],
                'led_status': ['label_led_status_title', 'label_led_status_value'],
                'cpu_usage': ['label_cpu_usage_title', 'label_cpu_usage_value'],
                'memory_usage': ['label_memory_usage_title', 'label_memory_usage_value'],
                'battery_status': ['label_dc_status_title', 'label_dc_status_value'],
                'interrupt_status': ['label_interrupt_status_title', 'label_interrupt_status_value']
            }
            
            # Update visibility for associated UI elements
            if monitor_name in monitor_ui_map:
                for ui_element_name in monitor_ui_map[monitor_name]:
                    if hasattr(self.window, ui_element_name):
                        ui_element = getattr(self.window, ui_element_name)
                        ui_element.setVisible(visible)
                        
                logger.debug(f"Updated UI visibility for {monitor_name}: {visible}")
                        
        except Exception as e:
            logger.error(f"Failed to update battery display visibility for {monitor_name}: {str(e)}")

    def _on_battery_refresh_clicked(self):
        """Handle battery refresh button click with pre-connection check"""
        logger.info("Battery refresh button clicked")
        
        # Add log entry
        self.log_manager.add_log_entry("INFO", "Battery refresh button clicked")
        
        try:
            # Check if battery monitor manager is available
            if not hasattr(self, 'battery_monitor_manager') or self.battery_monitor_manager is None:
                logger.info("Battery monitor not initialized, initializing now...")
                self._init_embedded_battery_monitor()
            
            # Check if monitoring is in progress
            if hasattr(self, 'battery_monitor_manager') and self.battery_monitor_manager:
                if self.battery_monitor_manager.is_monitoring:
                    logger.warning("Cannot refresh battery info while monitoring is active")
                    self.log_manager.add_log_entry("WARNING", "Cannot refresh battery info while monitoring is active")
                    return
                
                # Start refresh with pre-check
                self.execute_battery_refresh_with_pre_check()
            else:
                logger.error("Battery monitor manager not available after initialization")
                
        except Exception as e:
            logger.error(f"Error in battery refresh button click: {str(e)}")
            self.log_manager.add_log_entry("ERROR", f"Battery refresh error: {str(e)}")

    def execute_battery_refresh_with_pre_check(self):
        """Execute battery refresh with pre-check"""
        # Check if update is already in progress, avoid duplicate execution
        if hasattr(self, 'is_updating') and self.is_updating:
            logger.debug("System update already in progress, postponing battery refresh")
            return
            
        # Check if USB deployment is in progress, avoid command conflicts
        if hasattr(self, 'usb_deployment_in_progress') and self.usb_deployment_in_progress:
            logger.debug("USB deployment in progress, postponing battery refresh")
            return
        
        self.log_manager.add_log_entry("INFO", f"Checking connection before refreshing battery info for {self.device_id}...")
        
        # Show checking state
        original_text = self.window.pushButton_battery_refresh.text()
        self.window.pushButton_battery_refresh.setText("Checking...")
        self.window.pushButton_battery_refresh.setEnabled(False)
        
        # Use connection pre-check service to execute battery refresh
        self.connection_pre_check.execute_with_pre_check(
            device_id=self.device_id,
            operation_name="Battery Refresh",
            operation_callback=self._execute_battery_refresh,
            on_success=self._on_battery_refresh_pre_check_success,
            on_failure=lambda reason: self._on_battery_refresh_pre_check_failure(reason, original_text),
            check_timeout=12000  # 12 seconds timeout
        )
    
    def _execute_battery_refresh(self):
        """Execute the actual battery refresh operation"""
        # Get single battery reading
        if hasattr(self, 'battery_monitor_manager') and self.battery_monitor_manager:
            self.battery_monitor_manager.get_single_reading()
            
            # Restore refresh button state
            self.window.pushButton_battery_refresh.setText("Refresh Once")
            self.window.pushButton_battery_refresh.setEnabled(True)
    
    def _on_battery_refresh_pre_check_success(self):
        """Battery refresh pre-check success callback"""
        self.log_manager.add_log_entry("INFO", f"Connection verified, refreshing battery info for {self.device_id}")
    
    def _on_battery_refresh_pre_check_failure(self, reason: str, original_text: str):
        """Battery refresh pre-check failure callback"""
        self.log_manager.add_log_entry("ERROR", f"Connection check failed for battery refresh: {reason}")
        
        # Restore button to normal state since pre-check failed
        self.window.pushButton_battery_refresh.setText(original_text)
        self.window.pushButton_battery_refresh.setEnabled(True)
        
        # Show error message
        msg_box = QMessageBox(self.window)
        msg_box.setWindowTitle("Connection Check Failed")
        msg_box.setText("Device connection failed, battery refresh is canceled.")
        msg_box.setInformativeText(f"Ensure the device is connected, back up existing test content, then close the main window and return to Device Manager to reconnect the device.")
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setStandardButtons(QMessageBox.Ok)
        
        # Apply dark style sheet
        msg_box.setStyleSheet(self._get_dark_message_box_style())
        
        msg_box.exec()

    def _init_battery_chart(self):
        """Initialize battery chart widget"""
        try:
            from gui.widgets.battery_chart_widget import BatteryChartWidget
            
            # Create chart widget
            self.battery_chart_widget = BatteryChartWidget()
            
            # Add chart widget to the container in UI
            chart_container = self.window.widget_battery_chart_container
            chart_layout = chart_container.layout()
            chart_layout.addWidget(self.battery_chart_widget)
            
            # Connect UI controls to chart widget
            self._connect_chart_controls()
            
            logger.info("Battery chart widget initialized")
            
        except Exception as e:
            logger.error(f"Failed to initialize battery chart widget: {str(e)}")
            # Don't raise exception, chart is optional feature
            self.battery_chart_widget = None
    
    def _connect_chart_controls(self):
        """Connect chart control UI elements to chart widget functionality"""
        try:
            if not self.battery_chart_widget:
                return
            
            # Connect checkboxes to chart visibility toggles
            if hasattr(self.window, 'checkBox_show_battery'):
                self.window.checkBox_show_battery.toggled.connect(
                    self.battery_chart_widget._on_battery_toggled
                )
            
            if hasattr(self.window, 'checkBox_show_voltage'):
                self.window.checkBox_show_voltage.toggled.connect(
                    self.battery_chart_widget._on_voltage_toggled
                )
            
            if hasattr(self.window, 'checkBox_show_current'):
                self.window.checkBox_show_current.toggled.connect(
                    self.battery_chart_widget._on_current_toggled
                )
            
            if hasattr(self.window, 'checkBox_show_temperature'):
                self.window.checkBox_show_temperature.toggled.connect(
                    self.battery_chart_widget._on_temperature_toggled
                )
            
            # Connect clear button to clear data function
            if hasattr(self.window, 'pushButton_clear_chart'):
                self.window.pushButton_clear_chart.clicked.connect(
                    self.battery_chart_widget.clear_data
                )
            
            # Connect save button to save chart function
            if hasattr(self.window, 'pushButton_save_chart'):
                self.window.pushButton_save_chart.clicked.connect(
                    self._on_save_chart_clicked
                )
            
            logger.info("Chart controls connected successfully")
            
        except Exception as e:
            logger.error(f"Failed to connect chart controls: {str(e)}")
    
    def _on_embedded_monitoring_started(self):
        """Handle embedded monitoring started"""
        logger.info("Embedded battery monitoring started")
        
        # Add Battery Monitor to the mutual exclusion mechanism
        self._on_test_execution_started("Battery Monitor")
        
        # Set button to stop monitoring state
        self.window.pushButton_battery_monitor.setText("Stop Monitoring")
        self.window.pushButton_battery_monitor.setEnabled(True)
        self.window.pushButton_battery_monitor.setStyleSheet("""
            QPushButton {
                background-color: #D32F2F;
                color: white;
                border: none;
                padding: 6px 15px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #F44336;
            }
        """)
        self.log_manager.add_log_entry("INFO", "Battery monitoring started")
    
    def _on_embedded_monitoring_completed(self):
        """Handle embedded monitoring completed"""
        logger.info("Embedded battery monitoring completed")
        
        # Remove Battery Monitor from the mutual exclusion mechanism
        self._on_test_execution_completed("Battery Monitor")
        
        # Restore button to normal state
        self.window.pushButton_battery_monitor.setText("Start Monitoring")
        self.window.pushButton_battery_monitor.setEnabled(True)
        self.window.pushButton_battery_monitor.setStyleSheet("""
            QPushButton {
                background-color: #0078D7;
                color: white;
                border: none;
                padding: 6px 15px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #1C97EA;
            }
        """)
        self.log_manager.add_log_entry("INFO", "Battery monitoring stopped")
    
    def _on_embedded_monitoring_error(self, error_message: str):
        """Handle embedded monitoring error"""
        logger.error(f"Embedded battery monitoring error: {error_message}")
        
        # Remove Battery Monitor from the mutual exclusion mechanism on error
        self._on_test_execution_aborted("Battery Monitor", f"Error: {error_message}")
        
        # Restore button to normal state on error
        self.window.pushButton_battery_monitor.setText("Start Monitoring")
        self.window.pushButton_battery_monitor.setEnabled(True)
        self.window.pushButton_battery_monitor.setStyleSheet("""
            QPushButton {
                background-color: #0078D7;
                color: white;
                border: none;
                padding: 6px 15px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #1C97EA;
            }
        """)
        self.log_manager.add_log_entry("ERROR", f"Battery Monitor error: {error_message}")
    
    def _on_save_chart_clicked(self):
        """Handle Save Chart button click"""
        logger.info("Save Chart button clicked")
        
        try:
            # Check if chart widget is available
            if not hasattr(self, 'battery_chart_widget') or self.battery_chart_widget is None:
                logger.warning("Battery chart widget not available")
                self.log_manager.add_log_entry("WARNING", "Battery chart widget not available")
                
                # Show warning message
                msg_box = QMessageBox(self.window)
                msg_box.setWindowTitle("Can't save chart")
                msg_box.setText("Battery chart not initialized")
                msg_box.setInformativeText("Please start battery monitoring first, generate chart data, then save the chart.")
                msg_box.setIcon(QMessageBox.Warning)
                msg_box.setStandardButtons(QMessageBox.Ok)
                msg_box.setStyleSheet(self._get_dark_message_box_style())
                msg_box.exec()
                return
            
            # Check if there's any data to save
            if len(self.battery_chart_widget.timestamps) == 0:
                logger.warning("No chart data to save")
                self.log_manager.add_log_entry("WARNING", "No chart data to save")
                
                # Show warning message
                msg_box = QMessageBox(self.window)
                msg_box.setWindowTitle("No data to save")
                msg_box.setText("No chart data to save")
                msg_box.setInformativeText("Please start battery monitoring first, generate chart data, then save the chart.")
                msg_box.setIcon(QMessageBox.Information)
                msg_box.setStandardButtons(QMessageBox.Ok)
                msg_box.setStyleSheet(self._get_dark_message_box_style())
                msg_box.exec()
                return
            
            # Save chart with dialog
            saved_path = self.battery_chart_widget.save_chart_with_dialog()
            
            if saved_path:
                self.log_manager.add_log_entry("INFO", f"Battery chart saved to: {saved_path}")
                logger.info(f"Battery chart saved successfully: {saved_path}")
            else:
                logger.info("Chart save cancelled by user or failed")
                
        except Exception as e:
            error_msg = f"Failed to save battery chart: {str(e)}"
            logger.error(error_msg)
            self.log_manager.add_log_entry("ERROR", error_msg)
            
            # Show error message
            msg_box = QMessageBox(self.window)
            msg_box.setWindowTitle("Save failed")
            msg_box.setText("Error saving battery chart")
            msg_box.setInformativeText(str(e))
            msg_box.setIcon(QMessageBox.Critical)
            msg_box.setStandardButtons(QMessageBox.Ok)
            msg_box.setStyleSheet(self._get_dark_message_box_style())
            msg_box.exec()

    def _set_window_properties(self):
        """Set window properties, ensure it is recognized as part of the main application"""
        try:
            # Get application instance
            app = QApplication.instance()
            if not app:
                logger.warning("Unable to get QApplication instance")
                return
            
            # Use application icon
            if app.windowIcon() and not self.window.windowIcon():
                self.window.setWindowIcon(app.windowIcon())
                
            # Set window title to include platform name if available
            self.window.setWindowTitle("Main Window")
            
            # Set window flags to ensure it is recognized as the main application window
            self.window.setWindowFlags(self.window.windowFlags() | Qt.Window)
            
            # ensure the window can be resized
            self.window.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            
            # remove the fixed size constraint (if any)
            if self.window.minimumSize().width() == self.window.maximumSize().width() or \
               self.window.minimumSize().height() == self.window.maximumSize().height():
                # reset the minimum and maximum size
                self.window.setMinimumSize(600, 500)  # set a reasonable minimum size
                self.window.setMaximumSize(16777215, 16777215)  # set the maximum size to a large value
            
            # Set taskbar related properties on Windows
            if sys.platform == 'win32':
                try:
                    # we don't directly call Win32 API here
                    # let the system naturally associate the window - rely on the application identifier set earlier
                    pass
                except Exception as e:
                    logger.warning(f"Unable to set Windows window properties: {e}")
            
            logger.debug(f"Window properties set: Main Window")
        except Exception as e:
            logger.warning(f"Error setting window properties: {e}")
    
    def _update_dashboard_title(self):
        """Update dashboard title label based on platform name"""
        try:
            # Check if the label_title widget exists
            if hasattr(self.window, 'label_title'):
                if self.platform_name:
                    # Update the title to include platform name
                    title_text = f"System Monitor - {self.platform_name}"
                    self.window.label_title.setText(title_text)
                    logger.debug(f"Updated dashboard title to: {title_text}")
                else:
                    # Keep original title if no platform name
                    self.window.label_title.setText("System Monitoring Dashboard")
                    logger.debug("Using default dashboard title")
            else:
                logger.warning("label_title widget not found in main window")
        except Exception as e:
            logger.warning(f"Error updating dashboard title: {e}")

    @Slot()
    def _on_all_diagnostics_completed(self):
        """handle the event of all diagnostics completed"""
        self.log_manager.add_log_entry("INFO", "All diagnostic tests completed")
        
        # Restore UI controls after diagnostic tests complete
        self._on_test_execution_completed("Auto Diagnostic")

    @Slot()
    def _on_all_tests_completed(self):
        """Handle the event of all functionality tests completed"""
        self.log_manager.add_log_entry("INFO", "All functionality tests completed")
        
        # Restore UI controls after functionality tests complete
        self._on_test_execution_completed("Functionality Test All")

    def _convert_response_for_display(self, response, criteria, command, step_desc):
        """
        Convert response data to a more readable format for export
        
        Args:
            response: Original response string
            criteria: Test criteria
            command: Executed command
            step_desc: Step description
            
        Returns:
            Converted response string that's more user-friendly
        """
        if not response or response.strip() == "":
            return ""
        
        try:
            response_str = str(response).strip()
            
            if "i2cget -f -y 1 0x57 0x1f" in command.lower():
                return self._convert_embedded_eeprom_response(response_str)
            
            if "i2cdump -f -y -r 0x00-0x7f 1 0x57" in command.lower():
                return self._convert_dump_embedded_eeprom_response(response_str)
            
            if "i2ctransfer -f -y 1 w2@0x54 0x00 0x83 r19" in command.lower():
                return self._convert_eeprom_1_read_response(response_str)
            
            if "hexdump -c /sys/devices/platform/axi/ff030000.i2c/i2c-1/1-0054/eeprom" in command.lower():
                return self._convert_hexdump_eeprom_1_response(response_str)

            # Handle throughput speed conversion for USB and eMMC tests
            if "emmc_throughput" in step_desc.lower() or "usb_throughput" in step_desc.lower():
                return self._convert_throughput_response(response_str, step_desc)
            
            # Handle backlight brightness and power conversion
            if "brightness" in command.lower() or "bl_power" in command.lower():
                return self._convert_backlight_response(response_str, command)
            
            # Handle power button event conversion
            if "button event" in step_desc.lower() and "key_power" in response_str.lower():
                return self._convert_power_button_response(response_str)
            
            # Handle camera GPIO value conversion
            if "validate gpio value" in step_desc.lower():
                return self._convert_camera_gpio_response(response_str)
            
            # Handle panel resolution - check before i2c check since it may contain hex values
            if "resolution" in step_desc.lower():
                return self._convert_resolution_response(response_str)
            
            # Handle panel ID conversion
            if "panel id" in step_desc.lower():
                return self._convert_panel_id_response(response_str)
            
            # Handle NOR flash size conversion
            if "nor flash" in step_desc.lower() or "flash size" in step_desc.lower():
                return self._convert_nor_flash_response(response_str)
            
            # Handle LED status conversion
            if "led" in step_desc.lower() and "status" in step_desc.lower():
                return self._convert_led_status_response(response_str, step_desc)
            
            if "battery status" in step_desc.lower():
                return self._convert_battery_status_response(response_str, step_desc)
            
            # Handle DC value conversion
            if "gpio133" in command.lower():
                return self._convert_dc_value_response(response_str, step_desc)
            
            if "lsusb" in command.lower():
                return self._convert_lsusb_response(response_str)
            
            if "lspci" in command.lower():
                return self._convert_lspci_response(response_str)
            
            # Handle i2c transfer responses (hex values like "0x00 0x11")
            # More specific check: command contains i2ctransfer OR response has structured hex format
            if "i2ctransfer" in command.lower() or (
                "0x" in response_str and 
                not any(keyword in response_str.lower() for keyword in ["input device", "vendor", "product", "version"])
            ):
                return self._convert_i2c_response(response_str, command)
            
            # Handle eMMC size conversion
            if "emmc" in step_desc.lower() or "storage" in step_desc.lower() or "size" in command.lower():
                return self._convert_storage_response(response_str)
            
            # Handle version information
            if "version" in step_desc.lower() and "u-boot" in step_desc.lower():
                return self._convert_uboot_version_response(response_str)
            
            # Handle synctime information
            if "sync time" in criteria.lower():
                return self._convert_sync_time_response(response_str)
            
            # Clean up response text
            return self._clean_response_text(response_str)
            
        except Exception as e:
            logger.warning(f"Error converting response: {str(e)}")
            return response_str

    def _convert_embedded_eeprom_response(self, response):
        """Convert embedded eeprom response to readable format"""
        try:
            if "0xaa" in response.lower():
                return "0xaa"
            else:
                return response
            
        except Exception as e:
            logger.warning(f"Error converting embedded eeprom response: {e}")
            return response
    
    def _convert_dump_embedded_eeprom_response(self, response):
        """Convert dump embedded eeprom response to readable format"""
        try:
            if "10: ff ff ff ff ff ff ff ff ff ff ff ff ff ff ff aa" in response.lower():
                return "dump log with 'aa' in expected location"
            else:
                return response
        except Exception as e:
            logger.warning(f"Error converting dump embedded eeprom response: {e}")
            return response

    def _convert_eeprom_1_read_response(self, response):
        """Convert eeprom 1 read response to readable format"""
        try:
            if "0x32 0x30 0x32 0x35 0x2f 0x30 0x33 0x2f 0x32 0x38 0x20 0x32 0x30 0x3a 0x32 0x38 0x3a 0x33 0x36" in response.lower():
                return "0x32 0x30 0x32 0x35 0x2f 0x30 0x33 0x2f 0x32 0x38 0x20 0x32 0x30 0x3a 0x32 0x38 0x3a 0x33 0x36 = 2025/03/28 20:28:36"
            else:
                return response
        except Exception as e:
            logger.warning(f"Error converting eeprom 1 read response: {e}")
            return response

    def _convert_hexdump_eeprom_1_response(self, response):
        """Convert hexdump eeprom 1 response to readable format"""
        try:
            if "2025/03/28 20" in response.lower():
                return "dump log with '2025/03/28 20:28:36' in expected location"
            else:
                return response
        except Exception as e: 
            logger.warning(f"Error converting hexdump eeprom 1 response: {e}")
            return response

    def _convert_i2c_response(self, response, command):
        """Convert i2c hex response to readable format"""
        try:
            # Extract hex values from response
            hex_values = []
            lines = response.split('\n')
            
            for line in lines:
                if '0x' in line:
                    line_hex = [x.strip() for x in line.split() if x.startswith('0x')]
                    hex_values.extend(line_hex)
            
            if not hex_values:
                return response
            
            if len(hex_values) >= 10:
                # Skip first two bytes (status/length bytes), use next 8 bytes for data
                data_hex_values = hex_values[2:10]  # Take bytes 3-10
                # Convert hex values to ASCII characters
                ascii_chars = []
                for hex_val in data_hex_values:
                    try:
                        char_code = int(hex_val, 16)
                        if 32 <= char_code <= 126:  # Printable ASCII range
                            ascii_chars.append(chr(char_code))
                        else:
                            ascii_chars.append('?')  # Replace non-printable chars
                    except (ValueError, OverflowError):
                        ascii_chars.append('?')  # Replace invalid chars
                
                # Join characters and remove trailing nulls/spaces
                combined_value = ''.join(ascii_chars).rstrip('\x00').rstrip()
            elif len(hex_values) >= 3:
                # Skip the first byte (status byte 0x02) and use the next 2 bytes as data
                high_byte = int(hex_values[1], 16)
                low_byte = int(hex_values[2], 16)
                combined_value = (high_byte << 8) + low_byte

                # --- signed 16-bit conversion ---
                if combined_value & 0x8000:
                    combined_value -= 0x10000
                
                # high_byte = int(hex_values[1], 16)  # Second hex value
                # low_byte = int(hex_values[2], 16)   # Third hex value
                # combined_value = (high_byte << 8) + low_byte
            elif len(hex_values) == 2:
                # Two values: use both as data (high byte + low byte)
                high_byte = int(hex_values[0], 16)
                low_byte = int(hex_values[1], 16)
                combined_value = (high_byte << 8) + low_byte

                if combined_value & 0x8000:
                    combined_value -= 0x10000
            elif len(hex_values) == 1:
                # Single value
                combined_value = int(hex_values[0], 16)
            
            # Apply unit conversions based on step description
            if "0x51 0x00 0x19" in command.lower() or "0x81 0x00 0x15" in command.lower():
                return f"{round(combined_value, 2)}mV"
            elif "0x51 0x00 0x14" in command.lower() or "0x51 0x00 0x0a" in command.lower() or "0x81 0x00 0x14" in command.lower():
                return f"{round(combined_value/1000, 2)}A"
            elif "0x51 0x00 0x08" in command.lower():
                temp_celsius = round(combined_value/10 - 273.15, 2)
                return f"{temp_celsius}C"
            elif "0x51 0x00 0x18" in command.lower():
                return f"{combined_value}mAh"
            elif "0x51 0x00 0x0d" in command.lower():
                return f"{combined_value}%"
            elif "0x21 0x00 0x10" in command.lower():
                return f"v{combined_value}"
            elif "0x51 0x00 0x21" in command.lower():
                return f"{data_hex_values} = {combined_value}"
            else:
                return f"{combined_value} (decimal from hex: {' '.join(hex_values)})"
                
        except Exception as e:
            logger.warning(f"Error converting i2c response: {e}")
            return response

    def _convert_storage_response(self, response):
        """Convert storage size response to readable format"""
        try:
            lines = response.strip().split('\n')
            for line in lines:
                if line.strip().isdigit():
                    sectors = int(line.strip())
                    bytes_total = sectors * 512
                    gb_total = bytes_total / (1024 ** 3)
                    return f"{sectors} sectors = {gb_total:.2f}GB ({bytes_total:,} bytes)"
            return response
        except Exception as e:
            logger.warning(f"Error converting storage response: {e}")
            return response

    def _convert_power_button_response(self, response):
        """Convert power button event response to readable format"""
        try:
            import re
            
            # Look for power button events in the response
            # Pattern matches: type 1 (EV_KEY), code 116 (KEY_POWER), value X
            power_events = []
            
            # Split response into lines and process each line
            lines = response.split('\n')
            
            for line in lines:
                # Match power button events
                power_match = re.search(r'type 1 \(EV_KEY\), code 116 \(KEY_POWER\), value (\d)', line)
                if power_match:
                    value = power_match.group(1)
                    if value == "1":
                        power_events.append("type 1 (EV_KEY), code 116 (KEY_POWER), value 1: Pressed")
                    elif value == "0":
                        power_events.append("type 1 (EV_KEY), code 116 (KEY_POWER), value 0: Released")
            
            # If we found power button events, return them
            if power_events:
                # Join all events with newlines
                events_summary = '\n'.join(power_events)
                return f"Power button events detected:\n{events_summary}"
            
            # If no power button events found, return original response
            return response
            
        except Exception as e:
            logger.warning(f"Error converting power button response: {e}")
            return response
    
    def _convert_throughput_response(self, response, step_desc):
        """Convert throughput test response to extract read/write speed"""
        try:
            import re
            
            # Look for speed information in the response
            # Pattern matches formats like: "62.2 MB/s", "258 MB/s", "91.8 MB/s", "1.4 GB/s"
            speed_pattern = r'(\d+\.?\d*)\s+(MB/s|MiB/s|M/s|GB/s|GiB/s|G/s)'
            speed_matches = re.findall(speed_pattern, response)
            
            if not speed_matches:
                return response
            
            # Get the last speed value (usually the final transfer speed)
            final_speed_value = speed_matches[-1][0]
            final_speed_unit = speed_matches[-1][1]
            
            # Convert to MB/s for consistent display
            speed_value = float(final_speed_value)
            if final_speed_unit in ['GB/s', 'GiB/s', 'G/s']:
                speed_mb = speed_value * 1024  # convert GB to MB
                display_speed = f"{final_speed_value} {final_speed_unit} ({speed_mb:.1f} MB/s)"
            else:
                display_speed = f"{final_speed_value} {final_speed_unit}"
            
            # Determine if this is a read or write operation based on step description
            operation_type = "Unknown"
            if "write" in step_desc.lower():
                operation_type = "Write"
            elif "read" in step_desc.lower():
                operation_type = "Read"
            
            # Determine device type
            device_type = "Unknown"
            if "emmc_throughput" in step_desc.lower():
                device_type = "eMMC"
            elif "usb_throughput" in step_desc.lower():
                device_type = "USB"
            
            # Return formatted speed information
            return f"{device_type} {operation_type} Speed: {display_speed}"
            
        except Exception as e:
            logger.warning(f"Error converting throughput response: {e}")
            return response

    def _convert_backlight_response(self, response, command):
        """Convert backlight brightness and power response to readable format"""
        try:
            # Clean up the response to get the numeric value
            response_value = response.strip()
            
            # Handle brightness conversion
            if "brightness" in command.lower():
                # Convert brightness level (0-7) to percentage
                brightness_mapping = {
                    "0": "0%",
                    "1": "20%", 
                    "2": "30%",
                    "3": "40%",
                    "4": "50%",
                    "5": "60%",
                    "6": "80%",
                    "7": "100%"
                }
                
                if response_value in brightness_mapping:
                    return f"{response_value} = {brightness_mapping[response_value]}"
                else:
                    return f"{response_value} (Unknown brightness level)"
            
            # Handle bl_power conversion
            elif "bl_power" in command.lower():
                if response_value == "0":
                    return f"{response_value} = screen on"
                elif response_value == "1":
                    return f"{response_value} = screen off"
                else:
                    return f"{response_value} (Unknown power state)"
            
            return response
            
        except Exception as e:
            logger.warning(f"Error converting backlight response: {e}")
            return response

    def _convert_resolution_response(self, response):
        """Convert panel resolution response to readable format"""
        try:
            lines = response.split(" ")
            if lines[0] == "geometry":
                return f"Panel Resolution: {lines[1]}x{lines[2]} pixels"
            else:
                return response
        except Exception as e:
            logger.warning(f"Error converting resolution response: {e}")
            return response

    def _convert_panel_id_response(self, response):
        """Convert panel ID response to readable format with platform name"""
        try:
            panel_id = response.strip()
            
            # Panel ID to platform mapping based on panel_id_resolution_worker.py
            panel_id_mapping = {
                "01": "hydra_fhd (Note: argo's panel ID is same as hydra_fhd, use pic version to identify argo)",
                "00": "hydra", 
                "10": "gemini_fhd",
                "11": "gemini"
            }
            
            if panel_id in panel_id_mapping:
                platform_name = panel_id_mapping[panel_id]
                return f"{panel_id} {platform_name}"
            else:
                # Unknown panel ID, return as is
                return panel_id
                
        except Exception as e:
            logger.warning(f"Error converting panel ID response: {e}")
            return response

    def _convert_nor_flash_response(self, response):
        """Convert NOR flash size response to readable format"""
        try:
            # Extract hex size value from mtd response
            # Format: mtd0: 04000000 00020000 "MX29GL512G"
            lines = response.strip().split('\n')
            
            for line in lines:
                if 'mtd0:' in line:
                    parts = line.split()
                    if len(parts) >= 2:
                        # Get the first hex value after "mtd0:"
                        hex_size = parts[1].strip()
                        
                        # Convert hex to decimal bytes
                        size_bytes = int(hex_size, 16)
                        
                        # Convert to MB
                        size_mb = size_bytes / (1024 * 1024)
                        
                        return f"{hex_size} = {size_mb:.0f}MB ({size_bytes:,} bytes)"
            
            return response
        except Exception as e:
            logger.warning(f"Error converting NOR flash response: {e}")
            return response

    def _convert_led_status_response(self, response, step_desc):
        """Convert LED status response to readable format"""
        try:
            # LED status mapping from led_worker.py
            LED_STATUS_MAP = {
                0: "Off",       8: "Off",               16: "Off",      24: "Off",              32: "Off",
                1: "Blue",      9: "Blue Blinking",     17: "Blue",     25: "Blue Blinking",    33: "Blue",     49: "Blue Blinking",
                2: "Green",     10: "Green Blinking",   18: "Green",    26: "Green Blinking",   34: "Green",    50: "Green Blinking",
                3: "Cyan",      11: "Cyan Blinking",    19: "Cyan",     27: "Cyan Blinking",
                4: "Red",       12: "Red Blinking",     20: "Red",      28: "Red Blinking",     36: "Red",      52: "Red Blinking",
                5: "Fuchsia",   13: "Fuchsia Blinking", 21: "Fuchsia",  29: "Fuchsia Blinking",
                6: "Orange",    14: "Orange Blinking",  22: "Orange",   30: "Orange Blinking",  38: "Orange",   54: "Orange Blinking",
                7: "White",     15: "White Blinking",   23: "White",    31: "White Blinking",   40: "Yellow",   56: "Yellow Blinking",
            }
            
            # Extract hex values from response like i2c responses
            hex_values = []
            lines = response.split('\n')
            
            for line in lines:
                if '0x' in line:
                    line_hex = [x.strip() for x in line.split() if x.startswith('0x')]
                    hex_values.extend(line_hex)
            
            if not hex_values:
                return response
            
            # Only filter out status bytes if they appear at the first position
            data_hex = hex_values.copy()
            if len(data_hex) > 0 and data_hex[0] in ['0x02', '0x00']:
                data_hex = data_hex[1:]
            
            # Get the LED status value (typically the last hex value)
            if data_hex:
                led_status_value = int(data_hex[-1], 16)
            else:
                led_status_value = int(hex_values[-1], 16)
            
            # Look up LED status
            if led_status_value in LED_STATUS_MAP:
                led_status_text = LED_STATUS_MAP[led_status_value]
                return f"{led_status_value} = {led_status_text}"
            else:
                return f"{led_status_value} = Unknown LED Status"
                
        except Exception as e:
            logger.warning(f"Error converting LED status response: {e}")
            return response

    def _convert_battery_status_response(self, response, step_desc):
        """Convert battery status response to readable format"""
        try:
            # LED status mapping from led_worker.py
            BATTERY_STATUS_MAP = {
                128: "Charging",
                192: "Discharging",
                160: "Full Charged",
                224: "Full Charged",
                144: "Full Discharged",
                32770: "Initializing",
                32896: "Over Charged",
                16512: "Terminate Charge",
                16544: "Full Charged, Terminate Charge",
                20608: "Over Temperature, Terminate Charge",
                20672: "Over Temperature, Terminate Charge",
                4224: "Over Temperature - Charge",
                4288: "Over Temperature - Discharge",
                3008: "Remaining Capacity and Time Alarm, Terminate Discharge",
                2176: "Terminate Discharge",
                2432: "Remaining Time Alarm, Terminate Discharge",
                2688: "Remaining Capacity Alarm, Terminate Discharge",
                960: "Remaining Capacity and Time Alarm",
                704: "Remaining Capacity Alarm",
                448: "Remaining Time Alarm",
            }
            
            lines = response.strip().split('\n')
            hex_values = []
            
            for line in lines:
                # Skip command echo lines
                if 'i2ctransfer' in line or 'sleep' in line or 'root@' in line:
                    continue
                    
                # Look for hex values in the line
                if '0x' in line:
                    line_hex = [x.strip() for x in line.split() if x.startswith('0x')]
                    if line_hex:
                        hex_values.extend(line_hex)
            
            # Extract the correct hex values for battery commands
            # Typical i2c response format: 0x02 0xHH 0xLL (status + high byte + low byte)
            if len(hex_values) >= 3:
                # Skip the first byte (status byte 0x02) and use the next 2 bytes as data
                high_byte = int(hex_values[1], 16)  # Second hex value
                low_byte = int(hex_values[2], 16)   # Third hex value
                value = (high_byte << 8) + low_byte
            elif len(hex_values) == 2:
                # Two values: use both as data (high byte + low byte)
                high_byte = int(hex_values[0], 16)
                low_byte = int(hex_values[1], 16)
                value = (high_byte << 8) + low_byte
            elif len(hex_values) == 1:
                # Single value
                value = int(hex_values[0], 16)
            
            # Look up LED status
            if value in BATTERY_STATUS_MAP:
                battery_status_text = BATTERY_STATUS_MAP[value]
                return f"{value} = {battery_status_text}"
            else:
                return f"{value} = Unknown Battery Status"
                
        except Exception as e:
            logger.warning(f"Error converting LED status response: {e}")
            return response
    
    def _convert_dc_value_response(self, response, step_desc):
        """Convert DC value response to readable power status"""
        try:
            # DC value mapping
            DC_VALUE_MAP = {
                1: "Power On",
                0: "Power Off"
            }
            
            dc_value = int(response.strip())
            if dc_value in DC_VALUE_MAP:
                dc_status_text = DC_VALUE_MAP[dc_value]
                return f"{dc_value} = {dc_status_text}"
            else:
                return response
                
        except Exception as e:
            logger.warning(f"Error converting DC value response: {e}")
            return response

    def _convert_uboot_version_response(self, response):
        """Convert U-Boot version response to readable format"""
        try:
            pattern = r'U-Boot\s+([0-9]+\.[0-9]+[^\n]*?\([^)]+\))'
            match = re.search(pattern, response)
            
            if match:
                full_version = match.group(1).strip()
                return f"U-Boot version: {full_version}"
            
            return response
        except Exception as e:
            logger.warning(f"Error converting U-Boot version response: {e}")
            return response

    def _convert_camera_gpio_response(self, response):
        """Convert camera GPIO value response to camera model string"""
        try:
            # Parse GPIO value using the same logic as _parse_gpio_value
            lines = response.strip().split("\n")
            values = [line.strip() for line in lines if line.strip()]
            values.reverse()
            gpio_value_str = "".join(values)
            
            # Camera model mapping based on GPIO values
            camera_models = {
                "1001": "LVDS Titanium",
                "1110": "Scorpius", 
                "1101": "MIPI_VGA",
                "1011": "MIPI_720",
                "1010": "LVDS Smart cable",
                "0001": "Jig A",
                "0010": "Jig B"
            }
            
            # Look up the camera model
            if gpio_value_str in camera_models:
                camera_model = camera_models[gpio_value_str]
                return f"{gpio_value_str} = {camera_model}"
            else:
                return f"{gpio_value_str} (Unknown camera model)"
            
        except Exception as e:
            logger.warning(f"Error converting camera GPIO response: {e}")
            return response
    
    def _convert_lsusb_response(self, response):
        """Convert lsusb response to show device search result"""
        try:
            target_id = "1286:2046"
            target_id_2 = "1286:204e"
            lines = response.strip().split('\n')
            
            for line in lines:
                line = line.strip()
                if line and target_id in line:
                    return f"Device found: {line}"
                elif line and target_id_2 in line:
                    return f"Device found: {line}"
            
            return "Device not found"
            
        except Exception as e:
            logger.warning(f"Error converting lsusb response: {e}")
            return response
    
    def _convert_lspci_response(self, response):
        """Convert lspci response to show device search result"""
        try:
            target_name = "Marvell"
            lines = response.strip().split('\n')
            
            for line in lines:
                line = line.strip()
                if line and target_name in line:
                    return f"Device found: {line}"
            
            return "Device not found"
            
        except Exception as e:
            logger.warning(f"Error converting lspci response: {e}")
            return response
    
    def _convert_sync_time_response(self, response):
        """Convert sync time script output and ensure only the FIRST result is used."""
        try:
            lines = response.strip().split("\n")
            # ① 先找有無 Sync Time = PASS/FAIL
            output_lines = []
            for line in lines:
                output_lines.append(line)
                low = line.lower()

                if "sync time = pass" in low or "sync time = fail" in low:
                    return "\n".join(output_lines)   # 回傳第一段完整區塊
            # ② 如果沒有 PASS/FAIL → 找第一個 ntpdate 行
            for line in lines:
                if "ntpdate[" in line:
                    return line.strip()
            # ③ 若以上都沒有 → fallback 原始內容
            return response

        except Exception as e:
            logger.warning(f"Error converting sync time response: {e}")
            return response

    def _clean_response_text(self, response):
        """Clean up response text by removing command echoes, prompts, duplicate lines, and illegal XML characters."""
        try:
            # First, remove illegal XML characters that openpyxl cannot handle.
            # This regex removes most control characters except for tab, newline, and carriage return.
            if response and isinstance(response, str):
                response = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', response)
            else:
                return ""

            cleaned_lines = []
            seen_lines = set()
            lines = response.split('\n')
            
            for line in lines:
                line_stripped = line.strip()
                if not line_stripped:
                    continue
                    
                # Skip command echoes and prompts
                if any(skip_str in line_stripped for skip_str in [
                    "i2ctransfer", "grep", "uname", "free", "cat", "strings",
                    "root@", "#", "$", ">"
                ]):
                    continue
                
                # Remove duplicate lines - only keep the first occurrence
                if line_stripped not in seen_lines:
                    seen_lines.add(line_stripped)
                    cleaned_lines.append(line_stripped)
            
            return '\n'.join(cleaned_lines) if cleaned_lines else response
        except Exception as e:
            logger.warning(f"Error cleaning response text: {e}")
            return response if isinstance(response, str) else ""

    def _export_results_by_type(self, sheet, test_type, results, progress_records, exported_test_steps):
        """
        Export test results of a specific type to an openpyxl worksheet.
        
        Args:
            sheet: openpyxl worksheet object.
            test_type: 'diagnostic' or 'functionality'.
            results: Dictionary of test results.
            progress_records: Dictionary of test progress records (for functionality tests).
            exported_test_steps: Set to track exported steps to avoid duplicates.
        """
        data_exported = False
        
        all_test_ids = set(results.keys())
        if progress_records:
            all_test_ids.update(progress_records.keys())
        
        sorted_test_ids = sorted(list(all_test_ids))
        
        for test_id in sorted_test_ids:
            try:
                if test_type == "diagnostic" and not test_id.startswith("diagnostic_"):
                    continue
                if test_type == "functionality" and test_id.startswith("diagnostic_"):
                    continue

                step_templates = self.test_step_templates.get(test_type, {}).get(test_id, [])
                test_steps = results.get(test_id, {}).get("steps", [])
                
                if not step_templates and test_type == "functionality":
                    logger.warning(f"No step templates for functionality test {test_id}, skipping.")
                    continue

                if step_templates:
                    for template_index, template in enumerate(step_templates):
                        step_criteria = template.get('criteria', '')
                        if not step_criteria:
                            continue

                        step_desc = template.get('description', f'{test_type.title()} Test')
                        step_command = template.get('command', '')
                        is_manual_step = template.get('manual_only', False)
                        
                        step_message = "NOT_EXECUTED"
                        step_response = ""
                        step_time = "--:--:--"
                        
                        matching_step = next((s for s in test_steps if s.get('index') == template_index and s.get('description') == step_desc), None)

                        if matching_step:
                            step_message = matching_step.get('message', 'NOT_EXECUTED')
                            step_response = matching_step.get('response', '')
                            step_command = matching_step.get('command', step_command)
                            step_time = matching_step.get('time', step_time)
                        
                        if test_type == "functionality" and step_message == "NOT_EXECUTED":
                            worker = self.view_model.hardware_test_manager.test_workers.get(test_id)
                            if worker and template_index < len(worker.steps):
                                worker_step = worker.steps[template_index]
                                if hasattr(worker_step, 'passed') and worker_step.passed is not None:
                                    step_message = "PASS" if worker_step.passed else "FAIL"
                                    if is_manual_step:
                                        step_response = f"Manual interaction step - {step_message} (verified by user)"

                        if step_message == "NOT_EXECUTED":
                            continue

                        execution_signature = f"{test_id}_{step_desc}_{step_response}_{step_message}"
                        if execution_signature in exported_test_steps:
                            continue
                        exported_test_steps.add(execution_signature)

                        step_response = self._clean_response_text(step_response)
                        step_message = "SKIPPED" if isinstance(step_message, str) and "skip" in step_message.lower() else step_message
                        
                        response_converted = self._convert_response_for_display(step_response, step_criteria, step_command, step_desc)
                        
                        row_data = [
                            test_id, step_desc, step_criteria, step_message, step_command,
                            step_response, response_converted,
                            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), step_time
                        ]
                        sheet.append(row_data)
                        data_exported = True
                
                elif test_type == "diagnostic" and not step_templates and test_id in results:
                    result_data = results[test_id]
                    status = result_data.get("status", "")
                    message = result_data.get("details", {}).get("message", "")
                    if not message: continue
                    
                    row_data = [
                        test_id, "Diagnostic Test", "", f"{status}: {message}", "", "", "",
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), result_data.get("time", "--:--:--")
                    ]
                    sheet.append(row_data)
                    data_exported = True

            except Exception as ex:
                logger.warning(f"Error processing {test_type} result for {test_id}: {str(ex)}")
                continue
        
        return data_exported

    def _export_results(self):
        """Export test results to an XLSX file with optimized ordering"""
        try:
            if getattr(self, '_export_in_progress', False):
                return
            self._export_in_progress = True
            
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"test_results_{self.device_id}_{timestamp}.xlsx"
            
            file_path, _ = QFileDialog.getSaveFileName(
                self.window, "Export Test Results", default_filename, "Excel Files (*.xlsx)")
            
            if not file_path:
                self._export_in_progress = False
                return
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Test Results"
            
            sheet.append(["Tool Version", "v2.0.1_20251104"])
            sheet.append(["Config Version", "v2.0.1_20251104"])
            sheet.append(["Module", "Step", "Criteria", "Result", "Command", "Response", "Response_converted", "Timestamp", "Duration (sec)"])
            
            font_setting_list = [(1,1), (2,1), (3,1), (3,2), (3,3), (3,4), (3,5), (3,6), (3,7), (3,8), (3,9)]
            for row, col in font_setting_list:
                cell = sheet.cell(row, col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            exported_test_steps = set()
            data_exported = False
            
            # Export Diagnostic results
            logger.info("Exporting Auto Diagnostic results...")
            diagnostic_results = self.unified_test_results.get("diagnostic", {})
            if self._export_results_by_type(sheet, "diagnostic", diagnostic_results, {}, exported_test_steps):
                data_exported = True
            
            # Export Functionality Test results
            logger.info("Exporting Functionality Test results...")
            functionality_results = self.unified_test_results.get("functionality", {})
            functionality_progress = self.unified_test_progress.get("functionality", {})
            if self._export_results_by_type(sheet, "functionality", functionality_results, functionality_progress, exported_test_steps):
                data_exported = True
            
            # cell format setting should apply after data writed
            sheet.column_dimensions['A'].width = 25
            sheet.column_dimensions['B'].width = 40
            sheet.column_dimensions['C'].width = 48
            sheet.column_dimensions['D'].width = 8
            sheet.column_dimensions['E'].width = 55
            sheet.column_dimensions['F'].width = 69
            sheet.column_dimensions['G'].width = 55
            sheet.column_dimensions['H'].width = 18
            sheet.column_dimensions['I'].width = 15
            for row in range(4, 200):
                for col in range(1, 10):
                    cell = sheet.cell(row, col)
                    cell.alignment = Alignment(vertical='center', wrapText=True)
                    if col == 4:
                        result_value = str(cell.value).strip().upper()
                        if result_value == "PASS":
                            cell.font = Font(color="00B050", bold=True) # Green
                        elif result_value == "FAIL":
                            cell.font = Font(color="FF0000", bold=True) # Red
                        else:
                            cell.font = Font(bold=True) # Default

                        cell.alignment = Alignment(vertical='center', horizontal='center')
            sheet.freeze_panes = 'E4'
            
            # Export System Info to a new sheet if data is available
            if self.system_info_data:
                try:
                    info_sheet = workbook.create_sheet(title="System Info")
                    info_sheet.append(["Category", "Item", "Value"])
                    
                    # Style the header
                    for cell in info_sheet["1:1"]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Define the order of categories for a structured output
                    category_order = ['cpu', 'memory', 'storage', 'battery', 'firmware_os']
                    
                    # Flatten the nested dictionary and write to the sheet
                    for category in category_order:
                        if category in self.system_info_data:
                            info_dict = self.system_info_data[category]
                            if isinstance(info_dict, dict):
                                for key, value in info_dict.items():
                                    # Clean the value before writing
                                    display_value = self._clean_response_text(str(value))
                                    info_sheet.append([category.replace('_', ' ').title(), key, display_value])

                    # Auto-size columns for the new sheet for better readability
                    for column_cells in info_sheet.columns:
                        try:
                            length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                            info_sheet.column_dimensions[column_cells[0].column_letter].width = length + 2
                        except (ValueError, TypeError):
                            # Handle empty columns or other issues
                            pass
                    logger.info("System Info sheet created successfully.")
                except Exception as ex:
                    logger.error(f"Error creating System Info sheet: {str(ex)}")

            workbook.save(file_path)
            
            if data_exported:
                logger.info(f"Test results exported to: {file_path}")
            else:
                logger.warning("No data was exported to the XLSX file")
            
            self.clear_all_test_results()
            
            if hasattr(self.window, 'tableWidget_hardware_test_steps'):
                self.window.tableWidget_hardware_test_steps.setRowCount(0)
            if hasattr(self.window, 'progressBar_hardware_test'):
                self.window.progressBar_hardware_test.setValue(0)
                self.window.progressBar_hardware_test.setVisible(False)
            
            self.log_manager.add_log_entry("INFO", "Test and diagnostic records were cleared after exporting")
            
        except Exception as e:
            error_msg = f"Error exporting test results: {str(e)}"
            logger.error(error_msg)
            self.log_manager.add_log_entry("ERROR", error_msg)
        finally:
            self._export_in_progress = False

    @Slot(str, str, str)
    def _on_command_completed(self, device_id: str, command: str, response: str):
        """
        Handle command completed event
        
        Args:
            device_id: Device ID
            command: Command sent
            response: Command response
        """
        # Only process commands for this device
        if device_id != self.device_id:
            return
            
        # Log command and response - check if it is a test command
        # NOTE: this is a workaround to avoid logging the command and response for the test commands
        if not command.startswith("get_logs"):
            # check if the command has been logged
            if command not in self.logged_commands:
                self.log_manager.add_log_entry("INFO", f"[Command] {command}")
            # log the response in any case
            self.log_manager.add_log_entry("DEBUG", f"[Response] {response}")
            # remove the command from the tracking set
            self.logged_commands.discard(command)
        
        # Process response based on command
        if command.startswith("get_logs"):
            self._process_logs_response(response)
    
    def _update_dashboard(self):
        """Update dashboard information"""
        # use the connection pre-check to ensure the device connection is normal, instead of direct refresh
        # check if the update is already in progress, avoid duplicate execution
        if hasattr(self, 'is_updating') and self.is_updating:
            logger.debug("System info update already in progress, skipping dashboard update")
            return
            
        # Check if USB deployment is in progress, avoid command conflicts
        if hasattr(self, 'usb_deployment_in_progress') and self.usb_deployment_in_progress:
            logger.debug("USB deployment in progress, skipping dashboard update")
            return
            
        # use the connection pre-check to ensure the device connection is normal
        self._on_refresh_system_info()
    
    def _process_logs_response(self, response):
        """Process logs response"""
        # delegate to the log manager to process the logs response
        self.log_manager.process_logs_response(response)
    
    def show(self):
        """Show window and trigger USB package deployment followed by system info update"""
        # Check window properties again before showing, ensure it's part of the main application
        self._set_window_properties()
        
        # ensure the window can be resized
        from PySide6.QtCore import Qt
        self.window.setWindowFlags(
            self.window.windowFlags() | 
            Qt.Window | 
            Qt.WindowMinMaxButtonsHint | 
            Qt.WindowCloseButtonHint
        )
        self.window.setAttribute(Qt.WA_DeleteOnClose, False)  # prevent the window from being deleted when closed
        
        # show the window
        self.window.show()
        
        # ensure the window is raised and activated
        self.window.raise_()
        self.window.activateWindow()
        
        # Start USB package deployment first
        self._start_usb_package_deployment()
    
    def close(self):
        """Close the main window and cleanup resources"""
        try:
            logger.info(f"Closing main window for device {self.device_id}")
            
            # stop the waiting icon
            if hasattr(self, 'waiting_spinner') and self.waiting_spinner:
                self.waiting_spinner.stop()
                self.waiting_spinner = None
            
            # Clean up hardware test manager
            if self.view_model.hardware_test_manager:
                logger.debug("Cleaning up hardware test manager")
                # Stop any running tests
                self.view_model.hardware_test_manager.stop_current_test()
                if hasattr(self.view_model.hardware_test_manager, 'cleanup'):
                    self.view_model.hardware_test_manager.cleanup()
            
            # Clean up test manager
            if self.test_manager:
                logger.debug("Cleaning up test manager")
                self.test_manager.cleanup()
                self.test_manager = None
            
            # Clean up system info manager
            if self.system_info_manager:
                logger.debug("Cleaning up system info manager")
                # Disconnect signals
                try:
                    self.system_info_manager.info_update_completed.disconnect()
                    self.system_info_manager.info_update_error.disconnect()
                except Exception:
                    # Signals may already be disconnected
                    pass
                
                if hasattr(self.system_info_manager, 'cleanup'):
                    self.system_info_manager.cleanup()
                self.system_info_manager = None
            
            # Clean up HW/SW config manager
            if hasattr(self, 'hw_sw_config_manager') and self.hw_sw_config_manager:
                logger.debug("Cleaning up HW/SW config manager")
                try:
                    self.hw_sw_config_manager.config_updated.disconnect()
                except Exception:
                    pass
                self.hw_sw_config_manager = None
            
            # Clean up firmware & OS manager
            if hasattr(self, 'firmware_os_manager') and self.firmware_os_manager:
                logger.debug("Cleaning up firmware & OS manager")
                try:
                    self.firmware_os_manager.info_updated.disconnect()
                except Exception:
                    pass
                self.firmware_os_manager = None
            
            # Clean up log manager
            if self.log_manager:
                logger.debug("Cleaning up log manager")
                if hasattr(self.log_manager, 'cleanup'):
                    self.log_manager.cleanup()
                self.log_manager = None
            
            # Clean up auto diagnostic view
            if self.auto_diagnostic_view:
                logger.debug("Cleaning up auto diagnostic view")
                self.auto_diagnostic_view.cleanup()
                self.auto_diagnostic_view = None
            
            # Clean up connection pre-check service
            if hasattr(self, 'connection_pre_check') and self.connection_pre_check:
                logger.debug("Cleaning up connection pre-check service")
                self.connection_pre_check.cleanup()
                self.connection_pre_check = None
            
            # Clean up embedded battery monitor
            if hasattr(self, 'battery_monitor_manager') and self.battery_monitor_manager:
                logger.debug("Cleaning up battery monitor manager")
                try:
                    self.battery_monitor_manager.cleanup()
                except Exception:
                    pass
                self.battery_monitor_manager = None
            
            if hasattr(self, 'battery_service') and self.battery_service:
                logger.debug("Cleaning up battery service")
                try:
                    self.battery_service.cleanup()
                except Exception:
                    pass
                self.battery_service = None
            
            # Clean up battery chart widget
            if hasattr(self, 'battery_chart_widget') and self.battery_chart_widget:
                logger.debug("Cleaning up battery chart widget")
                try:
                    self.battery_chart_widget.cleanup()
                except Exception:
                    pass
                self.battery_chart_widget = None
            
            # Clean up CPU stress service
            if hasattr(self, 'cpu_stress_service') and self.cpu_stress_service:
                logger.debug("Cleaning up CPU stress service")
                try:
                    # Stop any running stress test
                    if self.cpu_stress_service.is_running:
                        self.cpu_stress_service.stop_stress_test()
                    # CPU stress service manages its own signal disconnection in cleanup()
                    # Cleanup service
                    self.cpu_stress_service.cleanup()
                except Exception:
                    pass
                self.cpu_stress_service = None
            
            # Clean up the view model
            if hasattr(self, 'view_model') and self.view_model and hasattr(self.view_model, 'cleanup'):
                logger.debug("Cleaning up view model")
                self.view_model.cleanup()
            
            # Disconnect signals
            try:
                # Disconnect view model signals
                if hasattr(self, 'view_model') and self.view_model:
                    try:
                        self.view_model.command_result.disconnect(self._on_command_completed)
                    except Exception:
                        pass  # If already disconnected, ignore the error
                
                # Disconnect test manager signals
                if hasattr(self, 'test_manager'):
                    try:
                        self.test_manager.all_tests_completed.disconnect(self._on_all_tests_completed)
                    except Exception:
                        pass  # Ignore error of already disconnected signal
            except Exception as e:
                logger.error(f"Error disconnecting signals: {e}")
            
            # Remove event filter
            if hasattr(self, 'window') and self.window:
                self.window.removeEventFilter(self)
            
            # Close the window
            if self.window:
                logger.debug("Closing the main window")
                self.window.close()
                self.window = None
            
            # Emit window closed signal
            self.window_closed.emit(self.device_id)
            
            logger.info(f"Main window resources cleaned up for device: {self.device_id}")
            
        except Exception as e:
            logger.error(f"Error during main window cleanup: {str(e)}")
            # Still try to close the window even if there was an error
            if self.window:
                self.window.close()
                self.window = None

    def set_ui_controls_state(self, enabled=True, exclude_widgets=None):
        """
        General UI control enable/disable method
        
        Args:
            enabled (bool): Whether to enable the controls
            exclude_widgets (list): List of widgets to exclude from the enable/disable operation
        """
        if exclude_widgets is None:
            exclude_widgets = []
        
        # handle the buttons
        for button in self.window.findChildren(QPushButton):
            if button not in exclude_widgets:
                button.setEnabled(enabled)
        
        # handle the combo boxes
        for combobox in self.window.findChildren(QComboBox):
            if combobox not in exclude_widgets:
                combobox.setEnabled(enabled)
        
        # handle the line edits
        for lineedit in self.window.findChildren(QLineEdit):
            if lineedit not in exclude_widgets:
                lineedit.setEnabled(enabled)
        
        # handle the tab widgets
        for tabwidget in self.window.findChildren(QTabWidget):
            if tabwidget not in exclude_widgets:
                for i in range(tabwidget.count()):
                    tabwidget.setTabEnabled(i, enabled)
        
        # Ensure the system log page's send command button and input box state is correct
        if hasattr(self.window, 'pushButton_send_command'):
            self.window.pushButton_send_command.setEnabled(enabled)
        if hasattr(self.window, 'lineEdit_command'):
            self.window.lineEdit_command.setEnabled(enabled)
                
        # if there is a TestManager, handle the test container buttons
        if hasattr(self, 'test_manager') and hasattr(self.test_manager, 'test_container'):
            self.test_manager.set_test_buttons_enabled(enabled)
            
        # if there is an AutoDiagnosticView, handle its buttons
        if hasattr(self, 'auto_diagnostic_view'):
            self.auto_diagnostic_view.set_buttons_enabled(enabled)

    def _on_tab_changed(self, index):
        """Handle tab change event"""
        # only save the current tab index when not updating
        if not self.is_updating:
            self.current_tab_index = index
            logger.debug(f"Tab changed to index {index}")

    def set_ui_controls_state_except_tabs(self, enabled=True):
        """
        Set UI controls state, but keep the tabs available
        
        Args:
            enabled (bool): Whether to enable the controls
        """
        # exclude the tab widgets
        exclude_widgets = []
        if hasattr(self.window, 'tabWidget'):
            exclude_widgets.append(self.window.tabWidget)
        
        # call the original method, but exclude the tab widgets
        self.set_ui_controls_state(enabled, exclude_widgets)

    def record_test_result(self, test_type, test_id, result_data):
        """record the test results to the unified storage"""
        if test_type in self.unified_test_results:
            self.unified_test_results[test_type][test_id] = result_data

    def record_test_progress(self, test_type, test_id, progress_data):
        """record the test progress to the unified storage"""
        if test_type in self.unified_test_progress:
            if test_id not in self.unified_test_progress[test_type]:
                self.unified_test_progress[test_type][test_id] = []
            self.unified_test_progress[test_type][test_id].append(progress_data)

    def clear_all_test_results(self):
        """clear all the test results"""
        for test_type in self.unified_test_results:
            self.unified_test_results[test_type].clear()
            self.unified_test_progress[test_type].clear()
        
        # 清理步骤模板
        for test_type in self.test_step_templates:
            self.test_step_templates[test_type].clear()
        
        # notify the views to reset the UI
        if hasattr(self, 'test_manager'):
            self.test_manager.reset_ui()
        if hasattr(self, 'auto_diagnostic_view'):
            self.auto_diagnostic_view.reset_ui()

    def mark_command_as_logged(self, command):
        """
        mark the command as logged, avoid duplicate logging
        
        Args:
            command: command string
        """
        self.logged_commands.add(command)
        # set a timer to clean up the outdated commands (3 minutes later)
        QTimer.singleShot(180000, lambda cmd=command: self.logged_commands.discard(cmd))

    @Slot(str)
    def _on_test_started(self, test_id: str):
        """Handle the event of test started, save step template information"""
        logger.info(f"Test started: {test_id}, saving step template")
        
        try:
            # get the step information from the active worker of the hardware test manager
            if hasattr(self.view_model.hardware_test_manager, 'active_test_worker') and self.view_model.hardware_test_manager.active_test_worker:
                worker = self.view_model.hardware_test_manager.active_test_worker
                logger.info(f"Found active worker for {test_id}: {type(worker).__name__}")
                
                if hasattr(worker, 'steps') and worker.steps:
                    logger.info(f"Worker has {len(worker.steps)} steps")
                    
                    # save the step template information
                    step_templates = []
                    for i, step in enumerate(worker.steps):
                        step_template = {
                            'index': i,
                            'description': getattr(step, 'description', ''),
                            'criteria': getattr(step, 'criteria', ''),
                            'command': getattr(step, 'command', ''),
                            'manual_only': getattr(step, 'manual_only', False),
                            'pre_condition': getattr(step, 'pre_condition', ''),
                            'post_check': getattr(step, 'post_check', ''),
                            'specification': getattr(step, 'specification', ''),
                        }
                        step_templates.append(step_template)
                        
                        # record the detailed information of each step
                        logger.debug(f"Step {i}: {step_template['description']} (criteria: {step_template['criteria']}, manual: {step_template['manual_only']})")
                    
                    # determine the test type
                    test_type = "functionality" if test_id.startswith("functionality_") else "diagnostic"
                    self.test_step_templates[test_type][test_id] = step_templates
                    
                    logger.info(f"Saved {len(step_templates)} step templates for {test_id}")
                    
                    # additional verification: check how many steps have criteria
                    steps_with_criteria = sum(1 for t in step_templates if t['criteria'])
                    logger.info(f"Steps with criteria: {steps_with_criteria}")
                    
                else:
                    logger.warning(f"No steps found in worker for {test_id}")
            else:
                logger.warning(f"No active worker found for {test_id}")
                
        except Exception as e:
            logger.error(f"Error saving step template for {test_id}: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
    
    def _get_dark_message_box_style(self):
        """return the dark style sheet for the message box"""
        return """
            QMessageBox {
                background-color: #2E2E2E;
                color: white;
            }
            QMessageBox QLabel {
                color: white;
                background-color: transparent;
            }
            QMessageBox QPushButton {
                background-color: #0078D7;
                color: white;
                border: none;
                padding: 6px 15px;
                border-radius: 3px;
                min-width: 60px;
            }
            QMessageBox QPushButton:hover {
                background-color: #1C97EA;
            }
            QMessageBox QPushButton:pressed {
                background-color: #00559F;
            }
            QMessageBox QPushButton:default {
                background-color: #0078D7;
                border: 2px solid #1C97EA;
            }
        """

    def _on_test_execution_started(self, test_type=""):
        """
        Handle UI state when test execution starts
        
        Args:
            test_type: Type of test being started (for logging)
        """
        logger.info(f"Test execution started: {test_type}")
        
        # Set test running flag
        self.is_test_running = True
        
        # Disable all UI controls except tabs and abort buttons
        exclude_widgets = []
        if hasattr(self.window, 'tabWidget'):
            exclude_widgets.append(self.window.tabWidget)
        
        # Keep abort buttons enabled
        if hasattr(self.window, 'button_abort_test'):
            exclude_widgets.append(self.window.button_abort_test)
        
        # If Battery Monitor is starting, keep the battery monitor button enabled so user can stop it
        if test_type == "Battery Monitor" and hasattr(self.window, 'pushButton_battery_monitor'):
            exclude_widgets.append(self.window.pushButton_battery_monitor)
        
        # If System Info Refresh is starting, keep the refresh button enabled for potential re-trigger
        if test_type == "System Info Refresh" and hasattr(self.window, 'pushButton_refresh'):
            exclude_widgets.append(self.window.pushButton_refresh)
        
        # Disable all other controls
        self.set_ui_controls_state(False, exclude_widgets)
        
        # Add log entry
        self.log_manager.add_log_entry("INFO", f"Test execution started - UI controls disabled")

    def _on_test_execution_completed(self, test_type=""):
        """
        Handle UI state when test execution completes
        
        Args:
            test_type: Type of test that completed (for logging)
        """
        logger.info(f"Test execution completed: {test_type}")
        
        # Clear test running flag
        self.is_test_running = False
        
        # Re-enable all UI controls
        self.set_ui_controls_state(True)
        
        # Add log entry
        self.log_manager.add_log_entry("INFO", f"Test execution completed - UI controls re-enabled")

    def _on_test_execution_aborted(self, test_type="", reason=""):
        """
        Handle UI state when test execution is aborted or fails
        
        Args:
            test_type: Type of test that was aborted (for logging)
            reason: Reason for abortion (for logging)
        """
        logger.info(f"Test execution aborted: {test_type} - {reason}")
        
        # Clear test running flag
        self.is_test_running = False
        
        # Re-enable all UI controls
        self.set_ui_controls_state(True)
        
        # Add log entry
        if reason:
            self.log_manager.add_log_entry("WARNING", f"Test execution aborted: {test_type} - {reason}")
        else:
            self.log_manager.add_log_entry("WARNING", f"Test execution aborted: {test_type}")

    def _on_individual_test_completed(self, test_id: str, test_type=""):
        """
        Handle UI state when an individual test completes (not part of a test sequence)
        
        Args:
            test_id: Test ID that completed
            test_type: Type of test (for logging)
        """
        # Only restore UI if this is an individual test (not part of Test All sequence)
        if hasattr(self, 'test_manager') and self.test_manager:
            # Check if Test All is currently running
            if not self.test_manager.is_test_all_running:
                logger.info(f"Individual test completed: {test_id} - restoring UI controls")
                
                # Clear test running flag
                self.is_test_running = False
                
                # Re-enable all UI controls
                self.set_ui_controls_state(True)
                
                # Add log entry
                self.log_manager.add_log_entry("INFO", f"Individual {test_type} test completed - UI controls re-enabled")
            else:
                logger.debug(f"Test {test_id} completed as part of Test All sequence - keeping UI disabled")
        else:
            # Fallback: restore UI if we can't determine test sequence status
            logger.info(f"Test {test_id} completed - restoring UI controls (fallback)")
            self.is_test_running = False
            self.set_ui_controls_state(True)
            self.log_manager.add_log_entry("INFO", f"{test_type} test completed - UI controls re-enabled")

    