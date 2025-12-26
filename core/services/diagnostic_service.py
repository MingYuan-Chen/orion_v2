import json
import os
import re
import time
from typing import Dict, List, Tuple, Any, Optional
from PySide6.QtCore import QObject, Signal, QTimer, Slot
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from core.models.serial_device_model import SerialDeviceModel
from util.logger import logger
from datetime import datetime
from util.command_loader import CommandLoader

class DiagnosticValidator:
    """
    Contains static methods for custom validation of diagnostic command outputs.
    """
    @staticmethod
    def validate_contains(output: str, expected: str) -> Tuple[bool, str]:
        """Simple validator to check if output contains expected string."""
        if expected in output:
            return True, f"Found expected output: {expected}"
        return False, f"All expected outputs not found"
    
    @staticmethod
    def validate_mac_address_pattern(output: str, **kwargs) -> Tuple[bool, str]:
        """
        Validate the mac address pattern by searching for it within the response.
        
        Args:
            mac_address: The response string containing the mac address to validate
        """
        # Use re.search to find the MAC address pattern anywhere in the response string.
        # The anchors (^) and ($) are removed to allow for surrounding text.
        mac_pattern = r'[0-9a-fA-F]{2}(:[0-9a-fA-F]{2}){5}'
        if re.search(mac_pattern, output):
            mac_address = re.search(mac_pattern, output).group()
            return True, f"Found valid mac address: {mac_address}"
        
        return False, f"The mac address is invalid"
    
    @staticmethod
    def validate_sync_time(output: str, platform_name: str, **kwargs) -> Tuple[bool, str]:
    # -----------------------
    # Odin platform
    # -----------------------
        if platform_name.lower() == "odin":
            server_ip = "192.168.6.11"
            # ntpdate success check
            ntp_ok = re.search(
                rf"(adjust|step)\s+time\s+server\s+{re.escape(server_ip)}\s+offset\b",
                output,
                flags=re.IGNORECASE
            ) is not None

            # date output (human readable)
            date_pattern = (
                r"\b\w{3}\s+\w{3}\s+\d{1,2}\s+"
                r"\d{2}:\d{2}:\d{2}\s+\w+\s+\d{4}\b"
            )
            date_matches = re.findall(date_pattern, output)

            # hwclock -r output
            rtc_pattern = (
                r"(?:"
                # ISO-like format
                r"\b\d{4}-\d{2}-\d{2}\s+"
                r"\d{2}:\d{2}:\d{2}"
                r"(?:\.\d+)?(?:[+-]\d{2}:\d{2})?\b"
                r"|"
                # BusyBox / date-like format
                r"\b\w{3}\s+\w{3}\s+\d{1,2}\s+"
                r"\d{2}:\d{2}:\d{2}\s+"
                r"(?:UTC\s+)?\d{4}\b"
                r")"
            )
            rtc_matches = re.findall(rtc_pattern, output)
            system_time_str = date_matches[-1] if date_matches else "N/A"
            rtc_time_str = rtc_matches[-1] if rtc_matches else "N/A"
            if ntp_ok :
                return True, (
                    f"NTP sync OK by time server {server_ip})-System Time:{system_time_str} / RTC Time:{rtc_time_str}"
                )
            if not ntp_ok:
                return False, (
                    f"NTP sync FAIL by {server_ip}-System Time:{system_time_str} / RTC Time:{rtc_time_str}"
                )

            return False, "Odin: validation failed"
        
        if platform_name.lower() == "athena":
            """Validator for Athena sync time."""
            pattern = r"(?:\w{3}\s\w{3}\s{1,2}\d{1,2}\s\d{2}:\d{2}:\d{2}\sUTC\s\d{4}|\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}(?:\.\d+)?(?:[+-]\d{2}:\d{2})?)"
            matches = re.findall(pattern, output)
            if matches and len(matches) == 3:
                hw_time = datetime.fromisoformat(matches[1]).replace(tzinfo=None)
                sw_time = datetime.strptime(matches[2], "%a %b %d %H:%M:%S %Z %Y").replace(tzinfo=None)
                time_difference = abs((hw_time - sw_time).total_seconds())
                if time_difference < 5:
                    return True, f"System time is auto calibrated from network"
                else:
                    return False, f"System time is not auto calibrated from network"

            return False, "No matched time string found"
    
    @staticmethod
    def validate_read_write(output: str, platform_name: str, key: str, **kwargs) -> Tuple[bool, str]:
        """Validator for USB/EMMC read write."""
        read_speed_threshold = 20
        write_speed_threshold = 30
        if platform_name.lower() == "odin":
            if key in ["diagnostic_USB-A_R/W-COM19(LEFT)", "diagnostic_USB-A_R/W-COM18(RIGHT)"]:
                read_speed_threshold = 20
                write_speed_threshold = 30
            elif key == "diagnostic_USB-C_R/W":
                read_speed_threshold = 45
                write_speed_threshold = 100
            elif key == "diagnostic_eMMc_R/W":
                read_speed_threshold = 50
                write_speed_threshold = 216
            elif key == "diagnostic_SD_card_R/W":
                read_speed_threshold = 45
                write_speed_threshold = 100
                
        elif platform_name.lower() == "athena":
            if key in ["diagnostic_USB1_R/W", "diagnostic_USB2_R/W"]:
                read_speed_threshold = 60
                write_speed_threshold = 60
            elif key == "diagnostic_eMMc_R/W":
                read_speed_threshold = 104
                write_speed_threshold = 104
        
        pattern = r'(\d+\.?\d*)\s+(MB/s|MiB/s|M/s|GB/s|GiB/s|G/s)'
        matches = re.findall(pattern, output)
        while len(matches) > 2: matches.pop(0)
        if matches and len(matches) == 2:
            write_speed_unit = matches[0][1]
            read_speed_unit = matches[1][1]

            if write_speed_unit in ['GB/s', 'GiB/s', 'G/s']:
                write_speed = float(matches[0][0]) * 1024
            else:
                write_speed = float(matches[0][0])

            if read_speed_unit in ['GB/s', 'GiB/s', 'G/s']:
                read_speed = float(matches[1][0]) * 1024
            else:
                read_speed = float(matches[1][0])

            device_type = "USB" if "USB" in key else "eMMC"
            if read_speed > read_speed_threshold and write_speed > write_speed_threshold:
                return True, f"{device_type} Read speed: {matches[1][0]} {matches[1][1]}, Write speed: {matches[0][0]} {matches[0][1]}"
            else:
                return False, f"{device_type} Read speed: {matches[1][0]} {matches[1][1]}, Write speed: {matches[0][0]} {matches[0][1]}"
        return False, "No matched speed string found"
    
    @staticmethod
    def validate_eeprom_for_odin(output: str, **kwargs) -> Tuple[bool, str]:
        # Normalize output
        output = output.upper()

        hex_matches = re.findall(r"[0-9A-F]{6,}", output)
        if not hex_matches:
            return False, "EEPROM FAIL — no valid HEX data found in output"
        signature = hex_matches[-1]
        if len(signature) < 6:
            return False, f"EEPROM FAIL — signature too short ({signature})"
        prefix = signature[:6]
        if prefix != "323232":
            return False, (
                f"Signature mismatch : "
                f"(got {prefix}"
            )

        return True, (
            f"Read signature={signature}, I2C bus=2, addr=0x4C"
        )
    @staticmethod
    def validate_eeprom_for_athena(output: str, **kwargs) -> Tuple[bool, str]:
        """Validator for EEPROM."""
        eeprom_1_byte = False
        eeprom_1_byte_dump = False
        eeprom_19_bytes = False
        eeprom_19_bytes_dump_1 = False
        eeprom_19_bytes_dump_2 = False
        responses = output.split("\n")
        for line in responses:
            line = line.strip()
            if line == "0xaa":
                eeprom_1_byte = True
            elif line == "0x32 0x30 0x32 0x35 0x2f 0x30 0x33 0x2f 0x32 0x38 0x20 0x32 0x30 0x3a 0x32 0x38 0x3a 0x33 0x36":
                eeprom_19_bytes = True
            elif line.startswith("10:") and "aa" in line:
                eeprom_1_byte_dump_1 = True
            elif line.startswith("00000080") and "2025/03/28 20" in line:
                eeprom_19_bytes_dump_1 = True
            elif line.startswith("00000090") and ":28:36" in line:
                eeprom_19_bytes_dump_2 = True
        if eeprom_1_byte and eeprom_19_bytes and eeprom_1_byte_dump_1 and eeprom_19_bytes_dump_1 and eeprom_19_bytes_dump_2:
            return True, "EEPROM Read and Write OK"
        return False, "EEPROM Read and Write Failed"
    
    @staticmethod
    def validate_charge_discharge_odin(output: str, **kwargs) -> Tuple[bool, str]:
        """
        Odin power validation (command-driven)
        - If command is Charge Status:
            current must be 0 ~ 3250 mA
            SoC 0~40% current must be 3000 ~ 3250 mA (Not yet)
            Soc 41%~80% current must be 1700~ 1600 mA (Not yet)
            SoC 81%~100% current must be 1650 ~ 0 mA (Not yet)
        - If command is Discharge Status:
            current must be -2000 ~ 0 mA
        """

        responses = output.split("\n")

        status = None
        soc = None
        current = None

        is_charge_cmd = False
        is_discharge_cmd = False

        CHARGE_STATUS = 128      # 0x00 0x80
        DISCHARGE_STATUS = 192   # 0x00 0xC0

        # -----------------------
        # Detect command intent
        # -----------------------
        for line in responses:
            if "Charge Status" in line:
                is_charge_cmd = True
            elif "Discharge Status" in line:
                is_discharge_cmd = True

        # Safety check
        if is_charge_cmd and is_discharge_cmd:
            return False, "Both Charge and Discharge commands detected"

        if not is_charge_cmd and not is_discharge_cmd:
            return False, "No Charge/Discharge command detected"
        # -----------------------
        # Parse values
        # -----------------------
        for line in responses:
            value = DiagnosticValidator._parse_battery_value(line)

            if not isinstance(value, int):
                continue

            # Status (display only)
            if value in (CHARGE_STATUS, DISCHARGE_STATUS):
                status = value
                continue

            # Battery SoC
            if 0 <= value <= 100:
                soc = value
                continue

            # Current (signed)
            if abs(value) >= 100:
                current = value

        if current is None:
            return False, "Current value not found"
        # -----------------------
        # PASS / FAIL by command
        # -----------------------
        if is_charge_cmd:
            if not (0 <= current <= 3250):
                return False, (
                    f"Now Charge current is {current}mA, "
                    f"Charge current should 0mA ~ 3120mA, "
                    f"Battery SoC: {soc}%"
                )
            mode = "Charging"

        else:  # Discharge command
            if not (-2000 <= current <= 0):
                return False, (
                    f"Now Current is {current}mA, "
                    f"Discharge current should -2000mA ~ 0A, "
                    f"Battery SoC: {soc}%"

                )
            mode = "Discharging"

        return True, (
            f"Battery SoC: {soc}%, "
            f"Current: {current}mA"
        )
    @staticmethod
    def validate_charge_for_athena(output: str, **kwargs) -> Tuple[bool, str]:
        """Validator for charge."""
        results = []
        responses = output.split("\n")
        for line in responses:
            result = DiagnosticValidator._parse_battery_value(line)
            if result:
                results.append(result)
        if len(results) == 3:
            if results[0] == "MD-BAT03":
                if results[1] == 768 and results[2] == 12592:
                    return True, f"Get available model name: {results[0]}, with correct charge current setting: {results[1]}mA, and charge voltage setting: {results[2]}mV"
                else:
                    return False, f"Incorrect charge current setting: {results[1]}mA or incorrect charge voltage setting: {results[2]}mV"
            elif results[0] is None:
                if results[1] == 192 and results[2] == 8992:
                    return True, f"Detected low battery mode, with correct charge current setting: 200mA, and charge voltage setting: 9000mV"
                else:
                    return False, f"Incorrect charge current setting: {results[1]}mA or incorrect charge voltage setting: {results[2]}mV"
        return False, "No matched charge values found"
    
    @staticmethod
    def _parse_battery_value(line) -> Any:
        """
        Parses the raw output based on the key.
        """

        try:
            if len(line) > 4 and line.startswith('0x'):
                line_hex = [x.replace('0x', '') for x in line.split() if x.startswith('0x')]

                # Battery model (ASCII)
                if len(line_hex) >= 4:
                    model = ""
                    for idx in range(1, len(line_hex)):
                        char_code = int(f"0x{line_hex[idx]}", 16)
                        if 32 <= char_code <= 126:
                            model += chr(char_code)
                    return model

                # 16-bit value (current / status / SoC)
                elif len(line_hex) == 2:
                    raw = int("0x" + line_hex[0] + line_hex[1], 16)

                    # 🔑 Convert to signed 16-bit
                    if raw >= 0x8000:
                        raw -= 0x10000

                    return raw

            return None

        except Exception as e:
            logger.error(f"Error parsing {line}: {e}")
            return f"Parse Error: {line}"
    @staticmethod
    def validate_power_button_for_odin(output: str, **kwargs) -> Tuple[bool, str]:
        """
        Odin Power Button validation (command-driven)
        - If 'PASS' appears in output: return True
        - If 'FAIL' appears in output: return False with timeout message
        """

        responses = output.splitlines()

        for line in responses:
            if "PASS" in line:
                return True, "Power button is detected"
            elif "FAIL" in line:
                return False, "Timeout: no key press detected within 10 seconds"

        return False, "No valid power button response found"

    @staticmethod
    def validate_probe_for_odin(output: str, **kwargs) -> Tuple[bool, str]:
        """
        Odin Probe validation

        PASS conditions:
        - 'Configuration complete!' exists
        - 'Planes saved: 14/14' exists
        - 'Check Sum Successful!' exists
        """

        if not output:
            return False, "Empty output"

        required_checks = {
            "Configuration complete!": "Console configuration failed",
            "Planes saved: 14/14": "Not all planes were saved successfully",
            "Check Sum Successful!": "CRC check failed",
        }

        missing = []

        for key, error_msg in required_checks.items():
            if key not in output:
                missing.append(error_msg)

        if missing:
            return False, "Probe test failed"

        return True, "Probe capture and CRC check successful"


class DiagnosticService(QObject):
    """
    Service to manage diagnostic execution and validation.
    """
    diagnostic_finished = Signal(str, bool, str) # key, success, message
    diagnostic_start = Signal(str) # key
    manual_check_requested = Signal(str, str) # key, message
    all_diagnostics_finished = Signal()

    def __init__(self, device_model: SerialDeviceModel, platform_name: str):
        super().__init__()
        self._model = device_model
        self._platform_name = platform_name
        self._diagnostics = {}
        self._running = False
        self._current_key = None
        self._current_commands = []
        self._current_output = []
        self._load_diagnostics()
        self._results_history = []

        # variables for Precondition setup
        self.dqa_package_path = None
        self.usb1_path = None
        self.usb2_path = None
        self.usb3_path = None
        self.eeprom_1_byte = None
        self.eeprom_19_bytes = None
        self.touch_qt_path = None
        self.test_audio_path = None
        
    def _load_diagnostics(self):
        # Load from resources/commands/{platform}/auto_diagnostic.json
        # Similar path logic as SystemInfoService/HWConfigService
        try:
            # Load commands dynamically
            mapper_folder_name = {
                "Athena": "athena",
                "Odin": "odin",
                "Gemini FHD": "gemini_fhd",
                "Gemini": "gemini",
                "Hydra FHD": "hydra_fhd",
                "Hydra": "hydra",
                "Argo": "argo"
            }
            _mapped_folder_name = mapper_folder_name.get(self._platform_name, "Unknown")
            self._diagnostics = CommandLoader.load_commands(_mapped_folder_name, "auto_diagnostic")
            if not self._diagnostics:
                logger.warning(f"No battery monitor commands found for platform: {self.platform_name}")
        except Exception as e:
            self._diagnostics = {}

    def run_diagnostics(self):
        """
        Runs all diagnostics sequentially.
        """
        self._precondition_setup()
        self._running = True
        self._queue = list(self._diagnostics.items())
        self._results_history = []
        self._run_next()

    def _run_next(self):
        if not self._running:
            return

        if not self._queue:
            self._save_to_excel()
            self.all_diagnostics_finished.emit()
            self._running = False
            return

        self._current_key, config = self._queue.pop(0)
        self._current_commands = list(config.get("commands", []))
        self._current_output = []
        
        if not self._current_commands:
            self._run_next()
            return

        # Emit start signal
        self.diagnostic_start.emit(self._current_key)

        # Start processing commands
        self._process_commands()

    def _process_commands(self):
        """
        Processes commands for the current diagnostic step.
        Supports pausing for manual checks.
        """
        while self._current_commands and self._running:
            cmd = self._current_commands.pop(0)

            if "manual_check_required" in cmd:
                # Get the message from the command
                message = cmd.replace("manual_check_required. ", "")
                # Pause execution and request user interaction
                self.manual_check_requested.emit(self._current_key, message)
                return

            if "sleep_required" in cmd:
                try:
                    duration = float(cmd.split(" ")[1])
                    output_str = f"Sleeping for {duration} second(s)"
                    self._current_output.append(output_str)
                    # Use QTimer to schedule the next step after the duration
                    # This yields control back to the event loop, keeping the UI responsive
                    QTimer.singleShot(int(duration * 1000), self._process_commands)
                    return
                except ValueError:
                    self._current_output.append(f"Invalid sleep duration in command: {cmd}")
                    continue

            # Execute command
            if "U-Boot" in cmd:
                response_lines = self._model.send_command_sync(cmd, timeout=20)
            elif "reboot" in cmd:
                response_lines = self._model.send_command_sync(cmd, wait_for="login:", timeout=60)
            elif "ts_test_mt -j 2 -v" in cmd:
                self._model.send_command_queued(cmd)
            elif "touch_qt_path" in cmd:
                if self.touch_qt_path:
                    cmd_replace = cmd.replace("touch_qt_path", self.touch_qt_path)
                    self._model.send_command_queued(f"'{cmd_replace}'")
            elif "usb1_path" in cmd:
                if self.usb1_path:
                    cmd = cmd.replace("usb1_path", self.usb1_path)
                    response_lines = self._model.send_command_sync(cmd)
            elif "usb2_path" in cmd:
                if self.usb2_path:
                    cmd_replace = cmd.replace("usb2_path", self.usb2_path)
                    response_lines = self._model.send_command_sync(cmd_replace)
            elif "usb3_path" in cmd:
                if self.usb3_path:
                    cmd_replace = cmd.replace("usb3_path", self.usb3_path)
                    response_lines = self._model.send_command_sync(cmd_replace)
            elif "eeprom_1_byte" in cmd:
                if self.eeprom_1_byte:
                    cmd_replace = cmd.replace("eeprom_1_byte", self.eeprom_1_byte)
                    response_lines = self._model.send_command_sync(cmd_replace)
            elif "eeprom_19_bytes" in cmd:
                if self.eeprom_19_bytes:
                    cmd_replace = cmd.replace("eeprom_19_bytes", self.eeprom_19_bytes)
                    response_lines = self._model.send_command_sync(cmd_replace)
            elif "test_audio_path" in cmd:
                if self.test_audio_path:
                    cmd_replace = cmd.replace("test_audio_path", self.test_audio_path)
                    response_lines = self._model.send_command_sync(cmd_replace, timeout=20)
            elif "ntpdate" in cmd:
                response_lines = self._model.send_command_sync(cmd, timeout=20)
            elif "aplay" in cmd:
                response_lines = self._model.send_command_sync(cmd, timeout=20)
            elif "odin_power_key_monitor.sh" in cmd:
                response_lines = self._model.send_command_sync(cmd, timeout=10)
            elif "configure_console.sh" or "probe-capture" or "check_crc.sh" in cmd:
                response_lines = self._model.send_command_sync(cmd, timeout=20)
            else:
                response_lines = self._model.send_command_sync(cmd)
            
            # Get output string
            if "touch_qt_path" in cmd or "ts_test_mt -j 2 -v" in cmd:
                output_str = "Touch Test Tool Launched"

            elif "usb1_path" in cmd and not self.usb1_path:
                output_str = "USB1 path not found"
            elif "usb2_path" in cmd and not self.usb2_path:
                output_str = "USB2 path not found"
            elif "usb3_path" in cmd and not self.usb3_path:
                output_str = "USB3 path not found"
            elif "eeprom_1_byte" in cmd and not self.eeprom_1_byte:
                output_str = "EEPROM 1 byte not found"
            elif "eeprom_19_bytes" in cmd and not self.eeprom_19_bytes:
                output_str = "EEPROM 19 bytes not found"
            else:
                output_str = "\n".join(response_lines)
            self._current_output.append(output_str)

        if not self._running:
            return

        # If we are here, either all commands are done or we stopped running
        if not self._current_commands:
            self._finish_current_diagnostic()

    @Slot(str)
    def resume_diagnostic(self, result: str):
        """
        Resumes the diagnostic process after a manual check.
        """
        if not self._running:
            return

        if result != "SKIP":
            self._current_output.append(f"manual_check_result_{result}")

        # Continue processing remaining commands
        self._process_commands()

    def _finish_current_diagnostic(self):
        combined_output = "\n".join(self._current_output)

        display_str = "Manual Check Result:"
        if self._current_key == "diagnostic_Backlight":
            display_str = "Brightness switch from 0 ~ 100% and turn on/off display:"
        elif self._current_key == "diagnostic_LCD":
            display_str = "LCD pattern switch to Red/Green/Blue/Black/White/Colorbar/gradient256/white frame/gray16,64,256:"
        elif self._current_key == "diagnostic_LED":
            display_str = "LED switch to Blue/Green/Red/Amber/Blink Blue/Blink Green/Blink Red/Blink Amber:"
        elif self._current_key == "diagnostic_Touch_9_Points":
            display_str = "Touch center/top/bottom/center-left/center-right/top-left/top-right/bottom-left/bottom-right:"
        elif self._current_key == "diagnostic_Touch_Drag_Draw":
            display_str = "Touch drag/draw:"
        elif self._current_key == "diagnostic_Camera_Preview":
            display_str = "LVDS, MIPI VGA, Scorpios, LVDS(smart cable) preview:"
        elif self._current_key == "diagnostic_HDMI_Mirror_Display":
            display_str = "HDMI mirror display:"
        elif self._current_key == "diagnostic_Audio_Record_Play":
            display_str = "Audio record from microphone and play by speaker:"
        # Validate
        if "manual_check_result_PASS" in combined_output:
            is_valid, msg = True, f"{display_str} Pass"
        elif "manual_check_result_FAIL" in combined_output:
            is_valid, msg = False, f"{display_str} Fail"
        else:
            is_valid, msg = self.validate_result(self._current_key, combined_output)
        
        self.diagnostic_finished.emit(self._current_key, is_valid, msg)
        
        # Store result
        self._results_history.append({
            "key": self._current_key,
            "success": is_valid,
            "message": msg,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        
        # Schedule next run
        self._run_next()

    def validate_result(self, key: str, output: str) -> Tuple[bool, str]:
        """
        Validates the output for a given diagnostic key.
        """
        try:
            config = self._diagnostics.get(key)
            if not config:
                return False, "Unknown diagnostic key"

            validate_func_name = config.get("validate_function")
            expected_response = config.get("expected_response")

            if validate_func_name:
                if hasattr(DiagnosticValidator, validate_func_name):
                    validator = getattr(DiagnosticValidator, validate_func_name)
                    params = {
                        "platform_name": self._platform_name,
                        "key": key,
                        "output": output
                    }
                    return validator(**params)
                else:
                    return False, f"Validator function '{validate_func_name}' not found"
            
            elif expected_response:
                # Default validation: check if output contains expected response(s)
                # expected_response can be a list or string
                if isinstance(expected_response, list):
                    result = False
                    target_string = ""
                    response_message = "No matched expected output found"
                    for expected in expected_response:
                        if expected in output:
                            result = True
                            target_string = expected
                            response_message = f"Found expected output: {target_string}"

                    if key == "diagnostic_CPU_Core_Type":
                        if result:
                            response_message = f"Expected CPU core type: {target_string}"
                        else:
                            response_message = "No matched CPU core type found"
                    elif key == "diagnostic_CPU_Processor":
                        if result:
                            response_message = f"Expected processor count: {target_string}"
                        else:
                            response_message = "No matched processor count found"
                    elif key == "diagnostic_eMMc_Size":
                        if result:
                            sectors = int(target_string)
                            # sector → bytes
                            bytes_total = sectors * 512
                            # bytes → GB (binary GiB but commonly called GB)
                            gb_total = bytes_total / (1024 ** 3)
                            display_str = f"{gb_total:.2f} GB"
                            response_message = f"Expected eMMC size: {display_str}"
                        else:
                            response_message = "No matched eMMC size found"
                    elif key == "diagnostic_Memory_Size":
                        if result:
                            response_message = f"Expected available memory size: {target_string} kB"
                        else:
                            response_message = "No matched available memory size found"
                    elif key == "diagnostic_PIC_Version":
                        if result:
                            version = int(target_string, 16)
                            response_message = f"Expected PIC version: v{version} ({target_string})"
                        else:
                            response_message = "No matched PIC version found"
                    elif key == "diagnostic_Console_Model":
                        if result:
                            response_message = f"Expected console model: {target_string}"
                        else:
                            response_message = "No matched console model found"
                    elif key == "diagnostic_Battery_Typical_Capacity":
                        if result:
                            display_str = DiagnosticValidator._parse_battery_value(target_string)
                            response_message = f"Expected battery typical capacity: {display_str}mAh"
                        else:
                            response_message = "No matched battery typical capacity found"
                    elif key == "diagnostic_Battery_Normal_Voltage":
                        if result:
                            voltage = DiagnosticValidator._parse_battery_value(target_string)
                            display_str = f"{voltage/1000:.1f}V"
                            response_message = f"Expected battery normal voltage: {display_str}"
                        else:
                            response_message = "No matched battery normal voltage found"
                    elif key == "diagnostic_UBoot_Version":
                        if result:
                            response_message = f"Expected U-Boot version: {target_string}"
                        else:
                            response_message = "No matched U-Boot version found"
                    elif key == "diagnostic_Kernel_Name":
                        if result:
                            response_message = f"Expected kernel name: {target_string}"
                        else:
                            response_message = "No matched kernel name found"
                    elif key == "diagnostic_Panel_Resolution":
                        if result:
                            display_str = target_string.replace(" ", "x")
                            response_message = f"Expected panel resolution: {display_str}"
                        else:
                            response_message = "No matched panel resolution found"
                    elif key == "diagnostic_Bluetooth_Controller":
                        if result:
                            response_message = "Expected Bluetooth controller found"
                        else:
                            response_message = "No matched Bluetooth controller found"
                    elif key == "diagnostic_WiFi_Controller":
                        if result:
                            response_message = "Expected WiFi controller found"
                        else:
                            response_message = "No matched WiFi controller found"
                    elif key == "diagnostic_Ethernet_Connection":
                        if result:
                            response_message = "Download google home page successfully via ethernet."
                        else:
                            response_message = "Download google home page failed via ethernet."
                    elif key == "diagnostic_WiFi_Connection":
                        if result:
                            response_message = "Download google home page successfully via wifi."
                        else:
                            response_message = "Download google home page failed via wifi."
                    return result, response_message
                else:
                    return DiagnosticValidator.validate_contains(output, str(expected_response))
            
            return True, "No validation criteria defined"
        except Exception as e:
            logger.error(f"Error validating diagnostic {key}: {e}")
            return False, str(e)
    
    def _precondition_setup(self):
        """
        Setup precondition for diagnostics.
        """
        self._find_original_eeprom_data()
        self._find_valid_usb_path()
   
    def _find_valid_usb_path(self) -> Tuple[bool, str]:
        
        
        if self._platform_name.lower() == "odin":
            response = "\n".join(
            self._model.send_command_sync("ls -l /dev/disk/by-path && ls /run/media")
            )

            # 對應你實際看到的 path 特徵
            USB_LEFT_KEY = "usb-0:1.2.2"
            USB_RIGHT_KEY = "usb-0:1.2.1"
            USB_C_KEY = "xhci-hcd.1.auto"

            usb_left_disk = None
            usb_right_disk = None
            usb_c_disk = None

            # -------- step 1: 找 sda / sdb / sdc --------
            for line in response.splitlines():
                if "-> ../../sd" not in line:
                    continue

                # 抓 sda / sdb / sdc
                disk = line.split("->")[-1].strip().replace("../../", "")

                if USB_LEFT_KEY in line:
                    usb_left_disk = disk
                elif USB_RIGHT_KEY in line:
                    usb_right_disk = disk
                elif USB_C_KEY in line:
                    usb_c_disk = disk

            # -------- step 2: 從 /run/media 找 partition --------
            media_parts = []
            for line in response.splitlines():
                if line.startswith("sda") or line.startswith("sdb") or line.startswith("sdc"):
                    media_parts = line.split()
                    break

            def find_mount(disk: str | None) -> str | None:
                if not disk:
                    return None
                for part in media_parts:
                    if part.startswith(disk):
                        return f"/run/media/{part}"
                return None

            self.usb1_path = find_mount(usb_left_disk)
            self.usb2_path = find_mount(usb_right_disk)
            self.usb3_path = find_mount(usb_c_disk)

            if not any([self.usb1_path, self.usb2_path, self.usb3_path]):
                return False, "找不到任何 USB mount 路徑"
            # -------- step 3: 找 dqa_package 並複製 --------
            self.dqa_package_path = None
            for usb_path in [self.usb1_path, self.usb2_path, self.usb3_path]:
                if not usb_path:
                    continue
                # 檢查 dqa_package 是否存在
                response = self._model.send_command_sync(
                    f"test -d {usb_path}/dqa_package && echo FOUND"
                )
                if response and "FOUND" in response[0]:
                    self.dqa_package_path = f"{usb_path}/dqa_package"
                    # 複製到 /root/dqa_package
                    self._model.send_command_sync("mkdir -p /home/root/dqa_package")
                    self._model.send_command_sync(
                        f"cp -r {self.dqa_package_path}/* /home/root/dqa_package/"
                    )
                    break
            if not self.dqa_package_path:
                return False, "USB mount found, but dqa_package not found"

            return True, (
                f"Odin USB paths: "
                f"left={self.usb1_path}, "
                f"right={self.usb2_path}, "
                f"typec={self.usb3_path} "
            )
            
        """
        Find valid usb path from 'ls -l /run/media'.
        e.g., 
        total 12
        drwxrwx---  3 root disk 4096 Jan  1  1970 'Main Data Partition-sdb1'
        drwxrwx--- 13 root disk 8192 Jan  1  1970  sda1
        """
        if self._platform_name.lower() == "athena":
            try:
                response = self._model.send_command_sync("ls -l /run/media")
                device_names = []
                for line in response:
                    line = line.strip()
                    if not line or line.startswith('total'):
                        continue
                    parts = line.split(None, 8)
                    if len(parts) == 9:
                        device_name = parts[8].strip("'")
                        device_names.append(device_name)

                # Prioritize assignment based on 'sda1' and 'sdb1'
                unassigned_names = []
                for name in device_names:
                    if 'sda1' in name and self.usb1_path is None:
                        self.usb1_path = f"/run/media/{name}"
                        file_path = f"{self.usb1_path}/dqa_package"
                        response = self._model.send_command_sync(f"ls {file_path}")
                        for line in response:
                            if "TouchTestQt64" in line and self.touch_qt_path is None:
                                self.touch_qt_path = f"{file_path}/TouchTestQt64"
                            if "odin_audio8k16S.wav" in line and self.test_audio_path is None:
                                self.test_audio_path = f"{file_path}/odin_audio8k16S.wav"
                    elif 'sdb1' in name and self.usb2_path is None:
                        self.usb2_path = f"/run/media/{name}"
                        file_path = f"{self.usb2_path}/dqa_package"
                        response = self._model.send_command_sync(f"ls {file_path}")
                        for line in response:
                            if "TouchTestQt64" in line and self.touch_qt_path is None:
                                self.touch_qt_path = f"{file_path}/TouchTestQt64"
                            if "odin_audio8k16S.wav" in line and self.test_audio_path is None:
                                self.test_audio_path = f"{file_path}/odin_audio8k16S.wav"
                    else:
                        logger.debug(f"Ignored device: {name}")
                
            except Exception as e:
                logger.error(f"Find valid usb path error: {str(e)}", exc_info=True)
                return False, f"Find valid usb path error: {str(e)}"

    def _find_original_eeprom_data(self):
        """
        Find original eeprom data from usb1, usb2, usb3.
        """
        try:
           response = self._model.send_command_sync("i2cget -f -y 1 0x57 0x1f")
           for line in response:
               if line.startswith("0x"):
                   self.eeprom_1_byte = line
           time.sleep(0.2)
           response = self._model.send_command_sync("i2ctransfer -f -y 1 w2@0x54 0x00 0x83 r19")
           for line in response:
               if line.startswith("0x"):
                   self.eeprom_19_bytes = line
           time.sleep(0.2)
        except Exception as e:
            logger.error(f"Find original eeprom data error: {str(e)}", exc_info=True)
            return False, f"Find original eeprom data error: {str(e)}"
    
    def _save_to_excel(self):
        """Saves the diagnostic results to an Excel file."""
        try:
            log_dir = os.path.join(os.getcwd(), "logs")
            if not os.path.exists(log_dir):
                os.makedirs(log_dir)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"diagnostic_report_{timestamp}.xlsx"
            filepath = os.path.join(log_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Diagnostic Report"
            
            # Headers
            headers = ["Timestamp", "Test Name", "Result", "Message"]
            ws.append(headers)
            
            # Style Headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for result in self._results_history:
                row = [
                    result["timestamp"],
                    result["key"].replace("diagnostic_", "").replace("_", " "),
                    "PASS" if result["success"] else "FAIL",
                    result["message"]
                ]
                ws.append(row)
                
                # Style Result Column
                result_cell = ws.cell(row=ws.max_row, column=3)
                if result["success"]:
                    result_cell.font = Font(color="008000", bold=True) # Green
                else:
                    result_cell.font = Font(color="FF0000", bold=True) # Red
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
                
            wb.save(filepath)
            logger.info(f"Diagnostic report saved to {filepath}")
            
        except Exception as e:
            logger.error(f"Failed to save diagnostic report: {e}")

    def disconnect(self):
        self._running = False
        if hasattr(self, '_queue'):
            self._queue.clear()
        if hasattr(self, '_model'):
            self._model.command_queue.clear()
