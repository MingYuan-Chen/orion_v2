import json
import os
import sys
import time
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional
from PySide6.QtCore import QObject, Signal, QTimer
from core.models.serial_device_model import SerialDeviceModel
from util.logger import logger
from util.command_loader import CommandLoader

class BatteryMonitorService(QObject):
    """
    Service to monitor battery status using commands defined in JSON.
    """

    mapper_folder_name = {
                "Athena": "athena",
                "Odin": "odin",
                "Gemini FHD": "gemini_fhd",
                "Gemini": "gemini",
                "Hydra FHD": "hydra_fhd",
                "Hydra": "hydra",
                "Argo": "argo"
            }

    LED_STATUS_MAP = {
        "athena": {
            1: "Blue", 33: "Blue",
            2: "Green", 34: "Green",
            4: "Red", 36: "Red",
            8: "Amber", 40: "Amber",
            17: "Blue Blink", 49: "Blue Blink",
            18: "Green Blink", 50: "Green Blink",  
            20: "Red Blink", 52: "Red Blink",
            24: "Amber Blink", 56: "Amber Blink",
            0: "Off", 32: "Off"
        },
        "odin": {
            1: "Blue", 17: "Blue",
            2: "Green", 18: "Green",
            6: "Yellow", 22: "Yellow",
            9: "Blink Blue", 25: "Blink Blue",
            10: "Blink Green", 26: "Blink Green",
            14: "Blink Yellow", 30: "Blink Yellow",
            0: "Off", 16: "Off"
        },
        "argo": {
            1: "Blue", 17: "Blue",
            2: "Green", 18: "Green",
            4: "Amber", 20: "Amber",
            9: "Blink Blue", 25: "Blink Blue",
            10: "Blink Green", 26: "Blink Green",
            12: "Blink Amber", 28: "Blink Amber",
            0: "Off", 16: "Off"
        },
        "other": {
            1: "Blue", 17: "Blue",
            2: "Green", 18: "Green",
            4: "Red", 20: "Red",
            9: "Blink Blue", 25: "Blink Blue",
            10: "Blink Green", 26: "Blink Green",
            12: "Blink Red", 28: "Blink Red",
            0: "Off", 16: "Off"
        }
    }

    INTERRUPT_STATUS_MAP = {
        0: "Normal",
        1: "No Battery",
        2: "Timeout",
        8: "Over Temperature - Charge",
        16: "Over Current - Charge",
        24: "Over Current & Temperature - Charge",
        32: "Over Temperature - Discharge",
        64: "Over Current - Discharge",
        96: "Over Current & Temperature - Discharge",
    }

    # [16 bit definitions]
    # 15: Overcharged, 14: Terminate Charge, 13: Undifined, 12: Overtemperature, 11: Terminate Discharge,
    # 10: Undifined, 9: Remaining Capacity, 8: Remaining Time, 7: Initialization, 6: Discharging,
    # 5: Full Charged, 4: Full Discharged, 0-3: Error Code
    BATTERY_STATUS_MAP = {
        128: "Charging",
        192: "Discharging",
        160: "Full Charged",
        224: "Full Charged",
        144: "Full Discharged",
        32770: "Initializing",
        32896: "Over Charged",
        16512: "Terminate Charge",
        16544: "Full Charged, Terminate Charge",
        20608: "Over Temperature, Terminate Charge",
        20672: "Over Temperature, Terminate Charge",
        4224: "Over Temperature - Charge",
        4288: "Over Temperature - Discharge",
        3024: "Terminate Discharge, Fully Discharged",
        3008: "Terminate Discharge, Remaining Capacity and Time Alarm",
        2688: "Terminate Discharge, Remaining Capacity Alarm",
        2432: "Terminate Discharge, Remaining Time Alarm",
        2176: "Terminate Discharge",
        960: "Remaining Capacity and Time Alarm",
        704: "Remaining Capacity Alarm",
        448: "Remaining Time Alarm",
    }

    SECONDS_PER_DAY = 86400
    SECONDS_PER_HOUR = 3600
    SECONDS_PER_MINUTE = 60

    # Signal to emit updated battery data
    battery_data_updated = Signal(dict)

    def __init__(self, device_model: SerialDeviceModel, platform_name: str = None):
        super().__init__()
        self._model = device_model
        self._platform_name = platform_name
        self._commands = {}
        self._last_results = {}
        self._unknown_count = 0
        self._load_commands()
        self._running = False
        
        # Excel Logging
        self._workbook = None
        self._worksheet = None
        self._excel_filename = None
        self._log_dir = os.path.join(os.getcwd(), "logs")
        if not os.path.exists(self._log_dir):
            os.makedirs(self._log_dir)
        self._monitoring_start_time = None
        
        # Timer for polling
        self._timer = QTimer()
        self._timer.setSingleShot(True)
        self._timer.timeout.connect(self.get_all_battery_info)
        self._interval_ms = 3000

    def start_monitoring(self, interval_ms: int = 3000):
        """Start periodic battery monitoring."""
        if not self._running:
            self._running = True
            self._interval_ms = interval_ms
            # Trigger immediately
            self.get_all_battery_info()

    def stop_monitoring(self):
        """Stop battery monitoring."""
        if self._running:
            self._running = False
            self._timer.stop()
            # Save and close Excel file
            if self._workbook:
                try:
                    self._workbook.save(self._excel_filename)
                    self._workbook.close()
                except Exception as e:
                    logger.error(f"Error saving Excel file on stop: {e}")
                self._workbook = None
                self._worksheet = None
                self._excel_filename = None

    def _load_commands(self):
        """
        Load battery monitor commands from JSON configuration.
        """
        try:
            # Load commands dynamically
            _mapped_folder_name = self.mapper_folder_name.get(self._platform_name, "Unknown")
            self._commands = CommandLoader.load_commands(_mapped_folder_name, "battery_monitor")
            if not self._commands:
                logger.warning(f"No battery monitor commands found for platform: {self.platform_name}")
        except Exception as e:
            self._commands = {}

    def get_all_battery_info(self) -> Dict[str, Any]:
        """
        Executes all configured commands and returns parsed results.
        """
        
        if not self._running:
            return

        start_time = time.time()
        if not self._monitoring_start_time:
            self._monitoring_start_time = start_time
        results = {}
        if not self._model.is_connected():
            logger.warning("Device not connected, cannot get battery info")
            # Even if not connected, we might want to retry later? 
            # For now, let's stop if disconnected or keep retrying?
            # Usually better to keep retrying or let the view model handle stop.
            # But to be safe, we schedule next run.
            if self._running:
                 self._timer.start(self._interval_ms)
            return results

        for key, cmd in self._commands.items():
            if not self._running: break # Check running status during loop
            try:
                response = self._model.send_command_sync(cmd[0])
                
                # Parse the value
                parsed_value = self._parse_value(key, response)

                # Optimization: Use last known good value if current is Unknown
                if "Unknown" in parsed_value and key in self._last_results:
                    if self._unknown_count < 3:
                        parsed_value = self._last_results[key]
                        self._unknown_count += 1
                    else:
                        parsed_value = "Unknown"
                elif parsed_value != "Unknown":
                    self._last_results[key] = parsed_value
                    self._unknown_count = 0

                results[key] = parsed_value
                
            except Exception as e:
                logger.error(f"Error getting info for {key}: {e}")
                results[key] = "Error"
            
        if self._running:
            timestamp = time.strftime("%m/%d %H:%M:%S")
            results["timestamp"] = timestamp
            end_time = time.time()

            total_seconds = int(end_time - self._monitoring_start_time)
            duration_days, remaining_seconds = divmod(total_seconds, self.SECONDS_PER_DAY)
            duration_hours, remaining_seconds = divmod(remaining_seconds, self.SECONDS_PER_HOUR)
            duration_minutes, duration_seconds = divmod(remaining_seconds, self.SECONDS_PER_MINUTE)
            duration_str = f"{duration_days}d {duration_hours:02d}:{duration_minutes:02d}:{duration_seconds:02d}"
            results["duration"] = duration_str
            # Save to Excel
            self._save_to_excel(results)

            self.battery_data_updated.emit(results)
            
            # Calculate elapsed time and adjust next interval
            
            elapsed_ms = int((end_time - start_time) * 1000)
            next_interval = max(0, self._interval_ms - elapsed_ms)
            
            # Schedule next run
            self._timer.start(next_interval)
            
        return results

    def _save_to_excel(self, data: Dict[str, Any]):
        """Appends data to an Excel file."""
        try:
            if not self._workbook:
                self._workbook = openpyxl.Workbook()
                self._worksheet = self._workbook.active
                self._worksheet.title = "Battery Log"
                
                filename = f"battery_log_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
                self._excel_filename = os.path.join(self._log_dir, filename)
                
                # Write Headers
                headers = [
                    "Timestamp", "Duration", "SoC", "Remaining Capacity", "Voltage", "Current",
                    "Temperature", "LED Status", "Battery Status", "Safety Status", "AC Present"
                ]
                self._worksheet.append(headers)
                
                # Style Headers
                for cell in self._worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                self._worksheet.freeze_panes = 'L2'
                
                # Initialize column widths
                self._column_widths = {}
                for i, header in enumerate(headers):
                    width = len(header) + 2
                    self._column_widths[i] = width
                    self._worksheet.column_dimensions[get_column_letter(i+1)].width = width

            timestamp = str(data.get("timestamp", "Unknown"))
            duration = str(data.get("duration", "Unknown"))
            voltage = str(data.get("voltage", "Unknown"))
            current = str(data.get("current", "Unknown"))
            rel_state = str(data.get("relative_state", "Unknown"))
            remaining_capacity = str(data.get("remaining_capacity", "Unknown"))
            temp = str(data.get("temperature", "Unknown"))
            batt_status = str(data.get("battery_status", "Unknown"))
            led_status = str(data.get("led_status", "Unknown"))
            interrupt = str(data.get("interrupt_status", "Unknown"))
            ac_present = str(data.get("ac_present", "Unknown"))
            
            # Prepare row data
            row_data = [
                timestamp, duration, rel_state, remaining_capacity, voltage, current,
                temp, led_status, batt_status, interrupt, ac_present
            ]

            self._worksheet.append(row_data)
            
            # Auto-adjust column widths
            for i, value in enumerate(row_data):
                width = len(value) + 2
                if width > self._column_widths.get(i, 0):
                    self._column_widths[i] = width
                    self._worksheet.column_dimensions[get_column_letter(i+1)].width = width

            # Save after every write to ensure data persistence
            self._workbook.save(self._excel_filename)
            
        except Exception as e:
            logger.error(f"Error saving to Excel: {e}")

    def _parse_value(self, key: str, response) -> Any:
        """
        Parses the raw output based on the key.
        TODO: Modify this method to implement specific parsing logic for your I2C values.
        """
        
        try:
            cpu_usage = None
            memory_usage = None
            for line in response:
                if len(line) > 4 and line.startswith('0x'):
                    line_hex = [x.replace('0x', '') for x in line.split() if x.startswith('0x')]
                    combined_hex = "0x" + line_hex[0] + line_hex[1]
                    hex_value = int(combined_hex, 16)

                    # Dispatch to specific parsers based on key
                    if key == "voltage":
                        if 0 <= hex_value <= 15000:
                            return f"{hex_value}mV ({combined_hex})"
                        else:
                            return "Unknown"
                    elif key == "current":
                        if hex_value > 32767:
                            signed_value = hex_value - 65536  # Convert to signed
                        else:
                            signed_value = hex_value
                        current = signed_value
                        if -5000 < current < 5000:
                            return f"{current}mA ({combined_hex})"
                        else:
                            return "Unknown"
                    elif key == "relative_state":
                        if 0 <= hex_value <= 100:
                            return f"{hex_value}% ({combined_hex})" # Battery percentage (0-100)
                        else:
                            return "Unknown"
                    elif key == "temperature":
                        temperature = round(float(hex_value/10)-273.2, 1)
                        if 0 < temperature < 120:
                            return f"{temperature}°C ({combined_hex})"  # Convert to Celsius
                        else:
                            return "Unknown"
                    elif key == "battery_status":
                        status = self.BATTERY_STATUS_MAP.get(hex_value, "Unknown")
                        return f"{status} ({combined_hex})"
                    elif key == "led_status":
                        mapped_name = self.mapper_folder_name.get(self._platform_name, "Unknown")
                        status = self.LED_STATUS_MAP.get(mapped_name, self.LED_STATUS_MAP.get("other")).get(hex_value, "Unknown")
                        return f"{status} ({combined_hex})"
                    elif key == "interrupt_status":
                        status = self.INTERRUPT_STATUS_MAP.get(hex_value, "Unknown")
                        return f"{status} ({combined_hex})"
                    elif key == "remaining_capacity":
                        return f"{hex_value}mAh ({combined_hex})"
                    elif key == "ac_present":
                        status = "Plugged" if hex_value == 55 else "Unplugged"
                        return f"{status} ({combined_hex})"
            # Default: return raw string stripped of whitespace
            return "Unknown"
            
        except Exception as e:
            logger.error(f"Error parsing {key}: {e}")
            return f"Parse Error: {response}"
